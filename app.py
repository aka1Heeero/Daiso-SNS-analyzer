import streamlit as st
import requests
import openpyxl
import re
import io
import time
import base64
import gspread
import pandas as pd
import altair as alt
import concurrent.futures
from datetime import datetime, date
from email.utils import parsedate_to_datetime
from google.oauth2.service_account import Credentials
try:
    from transformers import pipeline
    _HAS_TRANSFORMERS = True
except Exception:
    pipeline = None
    _HAS_TRANSFORMERS = False
from collections import Counter

# ============================
# 페이지 설정
# ============================
st.set_page_config(
    page_title="DAISO SNS ISSUE FINDER",
    page_icon="D",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================
# CSS
# ============================
from pathlib import Path

def _load_css(filename: str):
    """앱 파일 기준으로 CSS 파일을 찾아 주입.
    아래 후보 경로들을 순서대로 시도 (배포 구조 유연 대응):
    1. 앱 파일과 같은 폴더의 업로드해/
    2. 앱 파일 상위 폴더의 업로드해/
    3. 앱 파일과 같은 폴더 (평평한 구조)
    """
    here = Path(__file__).resolve().parent
    candidates = [
        here / "업로드해" / filename,
        here.parent / "업로드해" / filename,
        here / filename,
    ]
    for p in candidates:
        try:
            if p.exists():
                css = p.read_text(encoding="utf-8")
                st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)
                return
        except Exception:
            continue
    # CSS 없어도 앱은 뜨도록 조용히 넘김 (스타일만 빠짐)

_load_css("style.css")


# ============================================== 사용자 인증 (구글시트 기반)
def load_users_from_sheet():
    """구글시트 [users] 탭에서 사용자 목록 로드."""
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key("1iZS_bBlmZaMRFfW-l6XTP5zUZzit3vxhSIogEB-ynDM")
        ws = sh.worksheet("users")
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return {}
        headers = [h.strip().lower() for h in all_vals[0]]
        users = {}
        for row in all_vals[1:]:
            if len(row) < 4:
                continue
            r = dict(zip(headers, row))
            uid = r.get("id", "").strip()
            if uid:
                users[uid] = {
                    "password": str(r.get("password", "")).strip(),
                    "name": r.get("name", "").strip(),
                    "role": r.get("role", "user").strip(),
                }
        return users
    except Exception as e:
        st.error(f"users 시트 로드 실패 · {e}")
        return {}

def log_access(uid, name, action="login"):
    """구글시트 [access_log] 탭에 접속 기록 저장."""
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key("1iZS_bBlmZaMRFfW-l6XTP5zUZzit3vxhSIogEB-ynDM")
        ws = sh.worksheet("access_log")
        ws.append_row([uid, name, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    except Exception:
        pass

def change_password_in_sheet(uid, new_pw):
    """구글시트 [users] 탭에서 비밀번호 변경."""
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key("1iZS_bBlmZaMRFfW-l6XTP5zUZzit3vxhSIogEB-ynDM")
        ws = sh.worksheet("users")
        all_vals = ws.get_all_values()
        for idx, row in enumerate(all_vals[1:], 2):
            if row[0].strip() == uid:
                ws.update_cell(idx, 2, new_pw)
                return True
        return False
    except Exception:
        return False

def check_password():
    if st.session_state.get("authenticated"):
        return True

    if "_login_fail_cnt" not in st.session_state:
        st.session_state["_login_fail_cnt"] = 0
    if "_show_pw_change" not in st.session_state:
        st.session_state["_show_pw_change"] = False

    users = load_users_from_sheet()

    def _try_login():
        uid = st.session_state.get("login_id", "").strip()
        pw = st.session_state.get("login_pw", "").strip()
        if not uid or not pw:
            st.session_state["_login_empty"] = True
            return
        if uid in users and users[uid]["password"] == pw:
            st.session_state.authenticated = True
            st.session_state["current_user_id"] = uid
            st.session_state["current_user_name"] = users[uid]["name"]
            st.session_state["current_user_role"] = users[uid]["role"]
            st.session_state["_login_fail_cnt"] = 0
            if users[uid]["role"] == "admin":
                st.session_state["admin_mode"] = True
            log_access(uid, users[uid]["name"], "login")
        else:
            st.session_state["_login_fail_cnt"] = st.session_state.get("_login_fail_cnt", 0) + 1
            st.session_state["_login_error"] = True

    # ── 로그인 페이지 레이아웃 (좌: 브랜딩 2/3, 우: 폼 1/3) ──
    _logo_b64 = ""
    try:
        import os
        for p in ["assets/bi.jpg", "assets/BI.jpg", "BI.jpg", "bi.jpg"]:
            if os.path.exists(p):
                with open(p, "rb") as f:
                    _logo_b64 = base64.b64encode(f.read()).decode()
                break
    except Exception:
        pass
    _logo_html = f'<img src="data:image/jpeg;base64,{_logo_b64}" class="login-logo"/>' if _logo_b64 else '<div style="width:60px;height:60px;background:rgba(255,255,255,0.15);border-radius:50%;margin-bottom:1.5rem;"></div>'

    _load_css("login.css")

    left_col, right_col = st.columns([1.15, 1])
    with left_col:
        st.markdown(f"""
        <div class="login-left-brand">
          <div class="lb-inner">
            <div class="lb-top">
                {_logo_html}
                <div class="lb-brand">DAISO SNS ISSUE FINDER</div>
                <div class="lb-sub">고객 불만·호평 AI 분석 플랫폼</div>
            </div>
            <div class="lb-menu">
                <div class="lb-menu-item">
                    <div class="lb-mi-ic">01</div>
                    <div><div class="lb-mi-t">SNS 크롤링 수집</div><div class="lb-mi-d">블로그 · 카페 · 지식iN · 유튜브</div></div>
                </div>
                <div class="lb-menu-item">
                    <div class="lb-mi-ic">02</div>
                    <div><div class="lb-mi-t">AI 감성 분석</div><div class="lb-mi-d">긍정 · 부정 · 중립 자동 분류</div></div>
                </div>
                <div class="lb-menu-item">
                    <div class="lb-mi-ic">03</div>
                    <div><div class="lb-mi-t">품번 · 소분류 매칭</div><div class="lb-mi-d">상품 단위 이슈 도출</div></div>
                </div>
                <div class="lb-menu-item">
                    <div class="lb-mi-ic">04</div>
                    <div><div class="lb-mi-t">일자별 트렌드</div><div class="lb-mi-d">긍·부정 추이 시각화</div></div>
                </div>
            </div>
          </div>
          <div class="lb-credit">© 2026 데이터분석팀 · Internal Use Only</div>
        </div>
        """, unsafe_allow_html=True)
    with right_col:
        # 3번 실패 시 잠금
        if st.session_state["_login_fail_cnt"] >= 3:
            st.error("3회 로그인 실패. 관리자에게 문의하세요.")
            st.info("관리자 연락처 · 데이터분석팀")
            return False

        # 입력칸을 좁게 만들기 위한 내부 컬럼
        _pad_l, _form, _pad_r = st.columns([0.3, 2, 0.3])
        with _form:
            st.markdown('<span class="login-card-marker"></span>', unsafe_allow_html=True)
            if not st.session_state["_show_pw_change"]:
                # ── 로그인 화면 ──
                st.markdown("""
                <div class="login-card-head">
                    <div class="login-lock">•</div>
                    <div class="login-card-title">DAISO SNS ISSUE FINDER</div>
                    <div class="login-card-sub">고객 불만·호평 AI 분석 플랫폼</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown('<div class="login-flabel">아이디</div>', unsafe_allow_html=True)
                st.text_input("아이디", placeholder="아이디를 입력하세요", label_visibility="collapsed", key="login_id")
                st.markdown('<div class="login-flabel">비밀번호</div>', unsafe_allow_html=True)
                st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요", label_visibility="collapsed", key="login_pw", on_change=_try_login)
                if st.session_state.pop("_login_empty", False):
                    st.warning("아이디와 비밀번호를 입력하세요.")
                if st.session_state.pop("_login_error", False):
                    remain = 3 - st.session_state["_login_fail_cnt"]
                    st.error(f"아이디 또는 비밀번호가 올바르지 않습니다. (남은 시도: {remain}회)")
                if st.session_state.get("authenticated"):
                    st.rerun()
                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
                if st.button("로그인", use_container_width=True):
                    _try_login()
                st.markdown('<div class="login-divider"><span>또는</span></div>', unsafe_allow_html=True)
                if st.button("비밀번호 변경", use_container_width=True, type="secondary"):
                    st.session_state["_show_pw_change"] = True
                    st.rerun()
                st.markdown('<div class="login-foot">보안이 적용된 내부 시스템입니다.<br>© 2026 데이터분석팀. All rights reserved.</div>', unsafe_allow_html=True)

            elif st.session_state["_show_pw_change"]:
                # ── 비밀번호 변경 화면 ──
                st.markdown('<div style="font-size:0.95rem;font-weight:600;margin-bottom:0.5rem;">비밀번호 변경</div>', unsafe_allow_html=True)
                chg_id = st.text_input("", placeholder="아이디", label_visibility="collapsed", key="chg_id")
                chg_old = st.text_input("", type="password", placeholder="현재 비밀번호", label_visibility="collapsed", key="chg_old")
                chg_new = st.text_input("", type="password", placeholder="새 비밀번호", label_visibility="collapsed", key="chg_new")
                chg_new2 = st.text_input("", type="password", placeholder="새 비밀번호 확인", label_visibility="collapsed", key="chg_new2")
                if st.button("변경하기", use_container_width=True):
                    if not chg_id or not chg_old or not chg_new:
                        st.error("모든 항목을 입력해주세요.")
                    elif chg_new != chg_new2:
                        st.error("새 비밀번호가 일치하지 않습니다.")
                    elif chg_id not in users:
                        st.error("존재하지 않는 아이디입니다.")
                    elif users[chg_id]["password"] != chg_old:
                        st.error("현재 비밀번호가 틀렸습니다.")
                    else:
                        if change_password_in_sheet(chg_id.strip(), chg_new.strip()):
                            st.success("비밀번호가 변경되었습니다.")
                            st.session_state["_show_pw_change"] = False
                            st.rerun()
                        else:
                            st.error("변경 실패. 다시 시도해주세요.")
                if st.button("← 로그인으로 돌아가기", use_container_width=True, type="secondary"):
                    st.session_state["_show_pw_change"] = False
                    st.rerun()
    return False

if not check_password():
    st.stop()

# ============================================== API키
NAVER_CLIENT_ID     = st.secrets["NAVER_CLIENT_ID"]
NAVER_CLIENT_SECRET = st.secrets["NAVER_CLIENT_SECRET"]
YOUTUBE_API_KEY     = st.secrets.get("YOUTUBE_API_KEY", "")
GEMINI_API_KEY      = st.secrets.get("GEMINI_API_KEY", "")

# ============================================== 관리자 모드 세션 초기화
for _k, _v in {"admin_mode": False, "admin_exclude_kws": []}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ============================================== 구글시트 연동 (keyword / exclude_urls)
SHEET_ID = "1iZS_bBlmZaMRFfW-l6XTP5zUZzit3vxhSIogEB-ynDM"

def _get_gspread_client(readonly=True):
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"] if readonly else [
        "https://www.googleapis.com/auth/spreadsheets"
    ]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
    return gspread.authorize(creds)

@st.cache_data(ttl=600)
def load_keywords_from_sheet():
    """구글시트 [keyword] 탭에서 neg/pos/promo/exclude 키워드 로드."""
    try:
        gc = _get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("keywords")
        rows = ws.get_all_records()
        result = {"neg": [], "pos": [], "promo": [], "exclude": []}
        for r in rows:
            t = r.get("type", "").strip().lower()
            kw = (r.get("keyword", "") or r.get("keywords", "")).strip()
            if t in result and kw:
                result[t].append(kw)
        return result
    except Exception as e:
        st.warning(f"단어 시트 로드 실패 · {e}")
        return {"neg": [], "pos": [], "promo": [], "exclude": []}

@st.cache_data(ttl=600)
def load_excluded_urls_from_sheet():
    """구글시트 [exclude_urls] 탭에서 제외 URL 목록 로드."""
    try:
        gc = _get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("excluded_urls")
        rows = ws.get_all_records()
        return {r.get("url", "").strip() for r in rows if r.get("url", "").strip()}
    except Exception as e:
        st.warning(f"제외 링크 시트 로드 실패 · {e}")
        return set()

def append_keyword_to_sheet(kw_type, keywords):
    """구글시트 [keyword] 탭에 키워드 추가."""
    try:
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("keywords")
        ws.append_row([kw_type, keywords, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        load_keywords_from_sheet.clear()
    except Exception as e:
        st.error(f"시트 저장 실패: {e}")

def delete_keyword_from_sheet(kw_type, keyword):
    """구글시트 [keyword] 탭에서 키워드 삭제."""
    try:
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("keywords")
        records = ws.get_all_records()
        for idx, r in enumerate(records, 2):
            t = r.get("type", "").strip().lower()
            kw = (r.get("keyword", "") or r.get("keywords", "")).strip()
            if t == kw_type and kw == keyword:
                ws.delete_rows(idx)
                load_keywords_from_sheet.clear()
                return True
        return False
    except Exception as e:
        st.error(f"시트 삭제 실패: {e}")
        return False

def append_excluded_url_to_sheet(url, reason="관리자 제외"):
    """구글시트 [exclude_urls] 탭에 URL 추가."""
    try:
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("excluded_urls")
        ws.append_row([url, reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        load_excluded_urls_from_sheet.clear()
    except Exception as e:
        st.error(f"시트 저장 실패: {e}")


# ============================
# 골드셋 (감성 라벨링 데이터)
# ============================
@st.cache_data(ttl=600)
def load_goldset_from_sheet():
    """구글시트 [goldset] 탭에서 라벨링 데이터 로드.
    스키마: url | title | label | text | date | product_code | issue
    (product_code/issue는 없어도 됨, 없으면 빈 문자열로 채워짐)
    """
    try:
        gc = _get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("goldset")
        rows = ws.get_all_records()
        return rows
    except Exception:
        return []

def ensure_goldset_headers():
    """goldset 시트 헤더에 product_code, issue 컬럼이 없으면 자동 추가.
    앱 세션당 1회만 실행 (session_state 캐시).
    기존 데이터는 보존되며 새 컬럼은 빈 값으로 유지됨.
    """
    if st.session_state.get("_goldset_headers_ok"):
        return
    try:
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("goldset")
        headers = ws.row_values(1)
        # 기대 컬럼 (뒤에 추가할 것들)
        wanted = ["product_code", "issue"]
        to_add = [c for c in wanted if c not in headers]
        if to_add:
            new_headers = headers + to_add
            # 1행 갱신 (헤더만)
            end_col_letter = _col_num_to_letter(len(new_headers))
            ws.update(f"A1:{end_col_letter}1", [new_headers])
            load_goldset_from_sheet.clear()
        st.session_state["_goldset_headers_ok"] = True
    except Exception:
        # 실패해도 앱 동작에는 지장 없도록 조용히 넘김 (다음 세션에서 재시도)
        pass

def _col_num_to_letter(n: int) -> str:
    """1 -> A, 27 -> AA 형태의 시트 컬럼 문자 변환."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def append_goldset_to_sheet(url, title, label, text_snippet="", product_code="", issue=""):
    """구글시트 [goldset] 탭에 라벨링 데이터 추가.
    스키마: url | title | label | text | date | product_code | issue
    """
    try:
        ensure_goldset_headers()
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("goldset")
        ws.append_row([
            url,
            title,
            label,
            text_snippet[:500],  # 200→500 확장 (본문 근거 확보)
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            product_code,
            issue,
        ])
        load_goldset_from_sheet.clear()
    except Exception as e:
        st.error(f"골드셋 저장 실패: {e}")

def goldset_lookup_by_url():
    """정답셋을 URL 키로 조회 가능한 dict로 반환.
    같은 URL이 여러 번 라벨링된 경우 가장 최근 것을 사용.
    반환: {url: {label, product_code, product_name, category, issue, text, date}}
    """
    rows = load_goldset_from_sheet()
    if not rows:
        return {}
    # date 기준 오름차순 정렬 후 dict에 덮어써서 최신값이 남게 함
    def _key(r):
        return r.get("date", "") or ""
    rows_sorted = sorted(rows, key=_key)
    result = {}
    for r in rows_sorted:
        url = (r.get("url") or "").strip()
        if not url:
            continue
        code = str(r.get("product_code") or "").strip()
        # 품명DB에서 품명·카테고리 조인
        pname, cat = "", ""
        if code and not PRODUCT_DB.empty:
            row = PRODUCT_DB[PRODUCT_DB["품번"].astype(str).str.strip() == code]
            if not row.empty:
                pname = str(row.iloc[0].get("품명", "")).strip()
                cat = str(row.iloc[0].get("소분류", "")).strip()
        result[url] = {
            "label": (r.get("label") or "").strip(),
            "product_code": code,
            "product_name": pname,
            "category": cat,
            "issue": (r.get("issue") or "").strip(),
            "text": r.get("text") or "",
            "date": r.get("date") or "",
        }
    return result


def backfill_goldset_products_issues(progress_callback=None) -> dict:
    """정답셋의 기존 행 중 product_code/issue가 비어있는 것들을 Gemini로 자동 채움.
    - label=긍정/부정/중립인 것만 대상 (제외는 스킵)
    - 20개씩 배치 처리 (Gemini 1요청 = 20개)
    - Google Sheets batch_update로 두 컬럼만 한번에 업데이트
    - Gemini 하드리밋 초과 시 즉시 중단하고 지금까지 처리한 수 반환
    반환: {"processed": N, "updated": M, "skipped": S, "stopped_reason": str}
    """
    result = {"processed": 0, "updated": 0, "skipped": 0, "stopped_reason": ""}
    try:
        ensure_goldset_headers()
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet("goldset")
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            result["stopped_reason"] = "정답셋이 비어있음"
            return result
        headers = [h.strip() for h in all_vals[0]]
        # 필수 컬럼 인덱스
        idx_url = headers.index("url") if "url" in headers else 0
        idx_title = headers.index("title") if "title" in headers else 1
        idx_label = headers.index("label") if "label" in headers else 2
        idx_text = headers.index("text") if "text" in headers else 3
        idx_pcode = headers.index("product_code") if "product_code" in headers else None
        idx_issue = headers.index("issue") if "issue" in headers else None
        if idx_pcode is None or idx_issue is None:
            result["stopped_reason"] = "product_code / issue 컬럼이 없습니다 (시트 헤더 확인)"
            return result

        # 채워야 할 대상 rows 수집 (1-based row index)
        targets = []
        for ridx, row in enumerate(all_vals[1:], start=2):
            if len(row) <= max(idx_label, idx_text):
                continue
            label = (row[idx_label] if idx_label < len(row) else "").strip()
            if label not in ("긍정", "부정", "중립"):
                continue
            pcode = (row[idx_pcode] if idx_pcode < len(row) else "").strip()
            issue = (row[idx_issue] if idx_issue < len(row) else "").strip()
            if pcode and issue:
                continue  # 둘 다 이미 있으면 스킵
            title = (row[idx_title] if idx_title < len(row) else "").strip()
            text = (row[idx_text] if idx_text < len(row) else "").strip()
            targets.append({
                "row_index": ridx,
                "title": title,
                "text": text,
            })

        if not targets:
            result["stopped_reason"] = "채울 대상이 없습니다 (모두 이미 채워짐)"
            return result

        # 품명DB 후보 (상위 20개)
        candidates = []
        if not PRODUCT_DB.empty:
            for _, row in PRODUCT_DB.head(20).iterrows():
                candidates.append({
                    "품번": str(row.get("품번", "")).strip(),
                    "품명": str(row.get("품명", "")).strip(),
                    "소분류": str(row.get("소분류", "")).strip(),
                })

        BATCH = 20
        for i in range(0, len(targets), BATCH):
            # Gemini 호출 가능 여부 체크
            ok, reason = _gemini_can_call()
            if not ok:
                result["stopped_reason"] = f"AI 호출 중단 · {reason}"
                break
            chunk = targets[i:i + BATCH]
            items_for_ai = [
                {"idx": j, "title": t["title"], "description": t["text"]}
                for j, t in enumerate(chunk)
            ]
            ai_results = analyze_batch_with_gemini(items_for_ai, candidates)
            # idx로 매핑
            ai_by_idx = {}
            for r in ai_results:
                try:
                    aidx = int(r.get("idx"))
                except Exception:
                    continue
                ai_by_idx[aidx] = r

            # 배치 업데이트 준비
            batch_data = []
            for j, t in enumerate(chunk):
                result["processed"] += 1
                ai = ai_by_idx.get(j)
                if not ai:
                    result["skipped"] += 1
                    continue
                new_code = (ai.get("품번") or "").strip()
                new_issue = (ai.get("issue") or "").strip()
                if not new_code and not new_issue:
                    result["skipped"] += 1
                    continue
                # 시트 컬럼 좌표
                pcode_col = _col_num_to_letter(idx_pcode + 1)
                issue_col = _col_num_to_letter(idx_issue + 1)
                if new_code:
                    batch_data.append({"range": f"{pcode_col}{t['row_index']}", "values": [[new_code]]})
                if new_issue:
                    batch_data.append({"range": f"{issue_col}{t['row_index']}", "values": [[new_issue]]})
                result["updated"] += 1

            if batch_data:
                try:
                    ws.batch_update(batch_data)
                except Exception as e:
                    result["stopped_reason"] = f"시트 업데이트 실패 · {e}"
                    break

            if progress_callback:
                try:
                    progress_callback(min(i + BATCH, len(targets)), len(targets))
                except Exception:
                    pass

        # 캐시 무효화
        try:
            load_goldset_from_sheet.clear()
        except Exception:
            pass
        return result
    except Exception as e:
        result["stopped_reason"] = f"오류 · {e}"
        return result

def extract_goldset_keywords():
    """골드셋에서 긍정/부정 빈출 키워드 자동 추출."""
    goldset = load_goldset_from_sheet()
    if not goldset:
        return {"긍정": [], "부정": []}
    pos_texts = " ".join(r.get("text", "") for r in goldset if r.get("label") == "긍정")
    neg_texts = " ".join(r.get("text", "") for r in goldset if r.get("label") == "부정")
    # 2글자 이상 한글 단어 추출 후 빈도 계산
    pos_words = Counter(re.findall(r'[가-힣]{2,}', pos_texts))
    neg_words = Counter(re.findall(r'[가-힣]{2,}', neg_texts))
    # 상대방에 없고 3회 이상 등장한 단어만 추출
    pos_unique = [w for w, c in pos_words.most_common(30) if c >= 3 and neg_words.get(w, 0) <= 1]
    neg_unique = [w for w, c in neg_words.most_common(30) if c >= 3 and pos_words.get(w, 0) <= 1]
    return {"긍정": pos_unique[:15], "부정": neg_unique[:15]}


# 시트에서 키워드 로드
_sheet_kw = load_keywords_from_sheet()
EXCLUDED_URLS_FROM_SHEET = load_excluded_urls_from_sheet()


# ============================================== 구글시트 불러오기 (품번,품명,소분류)
@st.cache_data(ttl=3600)
def load_product_db():
    try:
        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
        )
        gc  = gspread.authorize(creds)
        sh  = gc.open_by_url(st.secrets["GSHEET_URL"])
        df  = pd.DataFrame(sh.sheet1.get_all_records())
        df.columns = [c.strip() for c in df.columns]
        if "품번" in df.columns:
            df["품번"] = df["품번"].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        return df
    except Exception as e:
        st.warning(f"품명 DB 로드 실패 · {e}")
        return pd.DataFrame(columns=["품번", "품명", "소분류"])

PRODUCT_DB = load_product_db()

VALID_PRODUCT_CODES = set()
if not PRODUCT_DB.empty and "품번" in PRODUCT_DB.columns:
    VALID_PRODUCT_CODES = set(PRODUCT_DB["품번"].dropna().astype(str).str.strip().tolist())

def load_subcategories():
    if not PRODUCT_DB.empty and "소분류" in PRODUCT_DB.columns:
        return list(PRODUCT_DB["소분류"].dropna().unique())
    return []

SUBCATEGORIES = load_subcategories()

# ── 제외할 소분류 (직접 수정) ──
EXCLUDE_SUBCATEGORIES = ["차", "자","커피","라면"]

# ============================================== AI모델링 (KLUE-RoBERTa + 룰베이스)
@st.cache_resource
def load_roberta():
    """감성분석 모델 로드. transformers 미설치 시 None 반환 (룰베이스만 사용)."""
    if not _HAS_TRANSFORMERS:
        return None
    model_name = "Chamsol/klue-roberta-sentiment-classification"
    # ONNX Runtime 최적화 시도 (2~3배 빠름)
    try:
        from optimum.onnxruntime import ORTModelForSequenceClassification
        from transformers import AutoTokenizer
        model = ORTModelForSequenceClassification.from_pretrained(model_name, export=True)
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        return pipeline("text-classification", model=model, tokenizer=tokenizer,
                        truncation=True, max_length=192, top_k=None, device=-1)
    except Exception:
        pass
    # fallback: 기본 PyTorch
    try:
        return pipeline("text-classification", model=model_name,
                        truncation=True, max_length=192, top_k=None, device=-1)
    except Exception:
        return None


# ============================================== Gemini AI (상품 + 이슈 자동 추출)
GEMINI_DAILY_LIMIT = int(st.secrets.get("GEMINI_DAILY_LIMIT", 100))
GEMINI_ENABLED = str(st.secrets.get("GEMINI_ENABLED", "true")).lower() != "false"

def _gemini_usage_read_sheet() -> int:
    """usage_log 시트에서 오늘 사용량 조회. 실패 시 0."""
    try:
        gc = _get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        try:
            ws = sh.worksheet("usage_log")
        except Exception:
            # 없으면 생성
            gc_rw = _get_gspread_client(readonly=False)
            sh_rw = gc_rw.open_by_key(SHEET_ID)
            ws = sh_rw.add_worksheet(title="usage_log", rows=1000, cols=3)
            ws.update("A1:B1", [["date", "count"]])
            return 0
        rows = ws.get_all_values()
        today = datetime.now().strftime("%Y-%m-%d")
        for r in rows[1:]:
            if len(r) >= 2 and r[0].strip() == today:
                try:
                    return int(r[1])
                except Exception:
                    return 0
        return 0
    except Exception:
        return 0

def _gemini_usage_inc_sheet(n: int = 1):
    """usage_log 시트의 오늘 카운트를 +n 원자적으로 증가.
    - 오늘 행 있으면 update, 없으면 append
    - 실패해도 앱 동작 지속 (하드리밋은 최선의 노력 방어)
    """
    try:
        gc = _get_gspread_client(readonly=False)
        sh = gc.open_by_key(SHEET_ID)
        try:
            ws = sh.worksheet("usage_log")
        except Exception:
            ws = sh.add_worksheet(title="usage_log", rows=1000, cols=3)
            ws.update("A1:B1", [["date", "count"]])
        rows = ws.get_all_values()
        today = datetime.now().strftime("%Y-%m-%d")
        for i, r in enumerate(rows[1:], start=2):
            if len(r) >= 2 and r[0].strip() == today:
                cur = 0
                try:
                    cur = int(r[1])
                except Exception:
                    pass
                ws.update_cell(i, 2, cur + n)
                return
        # 오늘 행 없음 → append
        ws.append_row([today, n])
    except Exception:
        pass

def _gemini_usage_today() -> int:
    """오늘 Gemini 호출 횟수 (모든 사용자 공유, 시트 저장).
    30초 세션 캐시로 성능·경합 밸런스.
    """
    now = datetime.now()
    cache_key = "_gemini_usage_cache"
    cache_ts_key = "_gemini_usage_cache_ts"
    ts = st.session_state.get(cache_ts_key)
    if ts and (now - ts).total_seconds() < 30:
        return st.session_state.get(cache_key, 0)
    val = _gemini_usage_read_sheet()
    st.session_state[cache_key] = val
    st.session_state[cache_ts_key] = now
    return val

def _gemini_usage_inc(n: int = 1):
    """호출 카운터 증가 (시트 + 세션 캐시 동시 갱신)."""
    _gemini_usage_inc_sheet(n)
    # 세션 캐시도 즉시 반영
    st.session_state["_gemini_usage_cache"] = st.session_state.get("_gemini_usage_cache", 0) + n
    st.session_state["_gemini_usage_cache_ts"] = datetime.now()

def _gemini_can_call() -> tuple:
    """호출 가능 여부 + 사유. 반환: (bool, reason)"""
    if not GEMINI_ENABLED:
        return False, "관리자에 의해 비활성화됨 (secrets: GEMINI_ENABLED=false)"
    if not GEMINI_API_KEY:
        return False, "API 키 미설정"
    used = _gemini_usage_today()
    if used >= GEMINI_DAILY_LIMIT:
        return False, f"오늘 사용량 한도 도달 ({used}/{GEMINI_DAILY_LIMIT}회)"
    return True, ""

@st.cache_resource
def _get_gemini_model():
    """Gemini 2.0 Flash 클라이언트를 세션 내 재사용."""
    if not GEMINI_API_KEY or not GEMINI_ENABLED:
        return None
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        return genai.GenerativeModel("gemini-2.0-flash")
    except Exception:
        return None

def _build_gemini_prompt(text: str, candidate_products: list) -> str:
    """분석 대상 본문 + 품명 후보 리스트로 프롬프트 구성.
    candidate_products: [{"품번","품명","소분류"}, ...]
    """
    # 후보 상품이 없으면 전체 DB에서 상위 200개만 (토큰 절약)
    if candidate_products:
        product_hint = "\n".join(
            f"- {p['품명']} (품번:{p['품번']}, 카테고리:{p.get('소분류','')})"
            for p in candidate_products[:20]
        )
        product_hint_block = f"\n[품명DB 후보]\n{product_hint}\n"
    else:
        product_hint_block = ""
    return f"""너는 다이소(DAISO) 고객 리뷰 분석 전문가다. 아래 SNS 글에서 다이소 상품에 대한 언급을 추출하라.

[분석할 글]
{text[:2500]}
{product_hint_block}
[출력 규칙]
- JSON 배열로만 답하라. 마크다운 코드블록 없이 순수 JSON만.
- 다이소 상품 언급이 없거나 광고/무관한 글이면 빈 배열 []을 반환.
- 각 항목 형식:
  {{"품명":"...", "품번":"...", "카테고리":"...", "label":"긍정|부정|중립", "issue":"핵심 불만/호평 한 줄"}}
- 품명DB 후보에 없는 상품이면 "품번"과 "카테고리"는 빈 문자열.
- issue는 15자 이내 한글 요약. 예: "뚜껑이 뻑뻑함", "가성비 좋음".
- 명확한 광고/홍보 글이면 [{{"label":"제외","issue":"광고"}}] 반환.
"""

def analyze_with_gemini(text: str, candidate_products: list = None) -> list:
    """Gemini로 상품·이슈·라벨 추출. 실패 시 빈 리스트.
    반환: [{"품명","품번","카테고리","label","issue"}, ...]
    """
    if not text or len(text.strip()) < 30:
        return []
    ok, _ = _gemini_can_call()
    if not ok:
        return []
    model = _get_gemini_model()
    if model is None:
        return []
    try:
        prompt = _build_gemini_prompt(text, candidate_products or [])
        # generation_config로 JSON 강제 & 온도 낮춤
        resp = model.generate_content(
            prompt,
            generation_config={
                "temperature": 0.2,
                "response_mime_type": "application/json",
                "max_output_tokens": 800,
            },
        )
        _gemini_usage_inc(1)
        raw = (resp.text or "").strip()
        if not raw:
            return []
        import json
        # 혹시 코드블록으로 감싸져 있으면 벗김
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip())
        data = json.loads(raw)
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            return []
        # 정제
        out = []
        for it in data:
            if not isinstance(it, dict):
                continue
            out.append({
                "품명": str(it.get("품명", "")).strip(),
                "품번": str(it.get("품번", "")).strip(),
                "카테고리": str(it.get("카테고리", "")).strip(),
                "label": str(it.get("label", "")).strip(),
                "issue": str(it.get("issue", "")).strip()[:60],
            })
        return out
    except Exception:
        return []

def gemini_available() -> bool:
    """UI에서 Gemini 사용 가능 여부 표시용."""
    ok, _ = _gemini_can_call()
    return ok and _get_gemini_model() is not None


def analyze_batch_with_gemini(items: list, candidate_products: list = None) -> list:
    """여러 결과를 한 요청으로 Gemini에 넘겨 초안 라벨 생성.
    items: [{"idx","title","description"}, ...] (idx는 원본 리스트 인덱스)
    반환: [{"idx", "품명","품번","카테고리","label","issue"}, ...] (실패 시 [])
    무료 티어 절약을 위해 최대 20개까지만 한 요청에 담음.
    """
    if not items:
        return []
    ok, _ = _gemini_can_call()
    if not ok:
        return []
    model = _get_gemini_model()
    if model is None:
        return []
    items = items[:20]
    # 후보 상품
    if candidate_products:
        product_hint = "\n".join(
            f"- {p['품명']} (품번:{p['품번']}, 카테고리:{p.get('소분류','')})"
            for p in candidate_products[:15]
        )
        product_hint_block = f"\n[품명DB 후보]\n{product_hint}\n"
    else:
        product_hint_block = ""
    # 배치 프롬프트
    lines = []
    for it in items:
        idx = it.get("idx")
        title = (it.get("title") or "").replace("\n", " ")[:120]
        desc = (it.get("description") or "").replace("\n", " ")[:600]
        lines.append(f"[{idx}] 제목: {title}\n본문: {desc}")
    joined = "\n\n".join(lines)
    prompt = f"""너는 다이소(DAISO) 고객 리뷰 분석 전문가다. 아래 {len(items)}개 글을 각각 분석하라.

[분석 대상]
{joined}
{product_hint_block}
[출력 규칙]
- JSON 배열로만 답하라. 마크다운 없이 순수 JSON.
- 각 항목 형식:
  {{"idx":N, "품명":"...", "품번":"...", "카테고리":"...", "label":"긍정|부정|중립|제외", "issue":"핵심 이슈 한 줄"}}
- idx는 반드시 원문 [숫자]와 동일하게 채워라.
- 다이소 상품 언급 없거나 광고면 label="제외", issue="광고" 또는 "무관".
- 품명DB 후보에 없으면 "품번"과 "카테고리"는 빈 문자열.
- issue는 15자 이내 한글 요약 (예: "뚜껑 뻑뻑함", "가성비 좋음").
"""
    try:
        resp = model.generate_content(
            prompt,
            generation_config={
                "temperature": 0.2,
                "response_mime_type": "application/json",
                "max_output_tokens": 4000,
            },
        )
        _gemini_usage_inc(1)
        raw = (resp.text or "").strip()
        if not raw:
            return []
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip())
        import json
        data = json.loads(raw)
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            return []
        out = []
        for it in data:
            if not isinstance(it, dict):
                continue
            out.append({
                "idx": it.get("idx"),
                "품명": str(it.get("품명", "")).strip(),
                "품번": str(it.get("품번", "")).strip(),
                "카테고리": str(it.get("카테고리", "")).strip(),
                "label": str(it.get("label", "")).strip(),
                "issue": str(it.get("issue", "")).strip()[:60],
            })
        return out
    except Exception:
        return []


# ============================================== 제외 학습 (label=제외 글의 특징어 추출)
def _extract_excluded_features(goldset_map: dict = None, min_count: int = 2, top_n: int = 30) -> list:
    """정답셋에서 label=제외로 라벨된 글들의 공통 특징어를 반환.
    - 유효(긍정/부정/중립) 글에는 드물고 제외 글에는 자주 등장하는 단어만 뽑음
    - TF-IDF 스타일 가중치 적용 (제외 비율 - 유효 비율)
    - 크롤링 파이프라인의 관련성 점수 계산에도 반영됨 (-5/-10점)
    """
    try:
        if goldset_map is None:
            goldset_map = goldset_lookup_by_url()
        if not goldset_map:
            return []
        excluded_texts, valid_texts = [], []
        for _, v in goldset_map.items():
            txt = (v.get("text") or "") + " " + (v.get("issue") or "")
            if v.get("label") == "제외":
                excluded_texts.append(txt)
            elif v.get("label") in ("긍정", "부정", "중립"):
                valid_texts.append(txt)
        if not excluded_texts:
            return []
        ex_words = Counter()
        for t in excluded_texts:
            for w in re.findall(r"[가-힣]{2,}", t):
                if w in _WC_STOPWORDS:
                    continue
                ex_words[w] += 1
        va_words = Counter()
        for t in valid_texts:
            for w in re.findall(r"[가-힣]{2,}", t):
                va_words[w] += 1
        n_ex = max(len(excluded_texts), 1)
        n_va = max(len(valid_texts), 1)
        scored = []
        for w, c in ex_words.most_common(300):
            if c < min_count:
                continue
            ex_rate = c / n_ex
            va_rate = va_words.get(w, 0) / n_va
            # 제외 비율이 유효 비율의 1.5배 이상 && 두 비율 차이가 0.05 이상
            if ex_rate >= 1.5 * (va_rate + 0.03) and (ex_rate - va_rate) >= 0.05:
                scored.append((w, ex_rate - va_rate, c))
        scored.sort(key=lambda x: (x[1], x[2]), reverse=True)
        return [w for w, _, _ in scored[:top_n]]
    except Exception:
        return []

def get_excluded_url_set() -> set:
    """excluded_urls 시트에 등록된 URL 집합 반환 (세션 캐시)."""
    if "_excluded_url_set" in st.session_state:
        return st.session_state["_excluded_url_set"]
    try:
        urls = load_excluded_urls_from_sheet() or []
    except Exception:
        urls = []
    s = set(u.strip() for u in urls if u)
    st.session_state["_excluded_url_set"] = s
    return s


# ============================================== 워드클라우드 (한글 폰트 자동 다운로드)
@st.cache_resource
def _get_korean_font_path() -> str:
    """Noto Sans KR 폰트를 캐시 폴더에 1회 다운로드 후 경로 반환.
    실패 시 시스템 기본 폰트 사용 (한글 깨질 수 있음)."""
    import os
    from pathlib import Path
    cache_dir = Path.home() / ".cache" / "sns_fonts"
    cache_dir.mkdir(parents=True, exist_ok=True)
    font_path = cache_dir / "NotoSansKR-Regular.ttf"
    if font_path.exists() and font_path.stat().st_size > 100000:
        return str(font_path)
    # Google Fonts 정적 파일 (TTF, 안정적)
    urls = [
        "https://github.com/google/fonts/raw/main/ofl/notosanskr/NotoSansKR%5Bwght%5D.ttf",
        "https://raw.githubusercontent.com/googlefonts/noto-cjk/main/Sans/OTF/Korean/NotoSansCJKkr-Regular.otf",
    ]
    for u in urls:
        try:
            r = requests.get(u, timeout=15)
            if r.status_code == 200 and len(r.content) > 100000:
                font_path.write_bytes(r.content)
                return str(font_path)
        except Exception:
            continue
    return ""

# 워드클라우드에서 제외할 흔한 단어 (SNS 텍스트 특성상 반복되는 단어)
_WC_STOPWORDS = {
    "다이소", "DAISO", "daiso", "네이버", "블로그", "카페", "유튜브", "youtube",
    "그리고", "그런데", "하지만", "그래서", "정말", "진짜", "완전", "너무", "약간",
    "이번", "저번", "요즘", "오늘", "어제", "내일", "지금", "이제", "다시", "그냥",
    "저는", "제가", "우리", "저희", "여러분", "친구", "가족", "사람", "사람들",
    "구매", "상품", "제품", "물건", "가격", "매장", "판매", "리뷰", "후기",
    "그거", "이거", "저거", "그것", "이것", "저것", "여기", "저기", "거기",
    "생각", "느낌", "정도", "부분", "경우", "때문", "이유", "얘기", "이야기",
    "있는", "없는", "있어", "없어", "있음", "없음", "있다", "없다", "이다", "된다",
    "합니다", "됩니다", "입니다", "했어요", "됐어요", "이에요", "예요", "이런", "저런", "그런",
}

def _tokenize_ko(text: str) -> list:
    """한글 명사 위주 토크나이저 (형태소 분석기 없이 간단히).
    - 2글자 이상 한글 단어 추출
    - 불용어 제거
    - 조사·어미 흔적 컷 (은/는/이/가/을/를/도/만/과/와/에 등)
    """
    if not text:
        return []
    # 한글만 남기고 공백 정리
    words = re.findall(r'[가-힣]{2,}', text)
    out = []
    for w in words:
        if w in _WC_STOPWORDS:
            continue
        # 흔한 조사 제거 (2글자면 스킵)
        if len(w) > 2:
            for suf in ["습니다", "입니다", "했어요", "이에요", "예요", "네요", "라고", "이라", "지만"]:
                if w.endswith(suf) and len(w) > len(suf) + 1:
                    w = w[:-len(suf)]
                    break
            for suf in ["에서", "부터", "까지", "으로", "이나", "라도"]:
                if w.endswith(suf) and len(w) > len(suf) + 1:
                    w = w[:-len(suf)]
                    break
            for suf in ["은", "는", "이", "가", "을", "를", "도", "만", "과", "와", "에", "의", "로"]:
                if w.endswith(suf) and len(w) > len(suf) + 2:
                    w = w[:-len(suf)]
                    break
        if len(w) >= 2 and w not in _WC_STOPWORDS:
            out.append(w)
    return out

def render_wordcloud(text: str, height: int = 300, max_words: int = 80, colormap: str = "Blues"):
    """텍스트를 받아 Streamlit에 워드클라우드 이미지를 렌더링.
    - 한글 폰트 자동 로드
    - wordcloud 라이브러리 미설치 시 빈도 상위 리스트로 대체
    """
    tokens = _tokenize_ko(text)
    if not tokens:
        st.info("워드클라우드로 표시할 단어가 부족합니다.")
        return
    freq = Counter(tokens)
    top = freq.most_common(max_words)
    try:
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt
        font_path = _get_korean_font_path()
        wc = WordCloud(
            font_path=font_path or None,
            width=1200, height=height * 2,
            background_color="white",
            colormap=colormap,
            max_words=max_words,
            prefer_horizontal=0.95,
            relative_scaling=0.4,
        ).generate_from_frequencies(dict(top))
        fig, ax = plt.subplots(figsize=(10, height / 60))
        ax.imshow(wc, interpolation="bilinear")
        ax.axis("off")
        plt.tight_layout(pad=0)
        st.pyplot(fig, clear_figure=True, use_container_width=True)
    except ImportError:
        # 라이브러리 없으면 텍스트 리스트로 대체
        html = '<div style="display:flex;flex-wrap:wrap;gap:0.4rem;">'
        maxc = top[0][1] if top else 1
        for w, c in top:
            size = 0.75 + (c / maxc) * 1.4
            weight = 700 if c / maxc > 0.5 else 500
            html += f'<span style="font-size:{size:.2f}rem;font-weight:{weight};color:#2563EB;padding:0.15rem 0.5rem;background:#EFF4FF;border-radius:4px;">{w}</span>'
        html += "</div>"
        st.markdown(html, unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"워드클라우드 생성 실패: {e}")



# ============================
# 룰베이스 & 앙상블
# ============================
NEGATIVE_PATTERNS = [
    r"불만|짜증|별로|최악|실망|환불|불량|교환|형편없|쓰레기|구려",
    r"고장|터졌|망가|깨졌|불편|아쉬워|위험|조심|주의|문제|하자",
    r"뜯겨|냄새|오염|불결|지저분|더럽|싸구려|허접|대충",
    r"클레임|환급|반품",
    r"재구매\s*안|비추|속았|낚였|사기|뻥|가짜",
    r"별점\s*1|별\s*1|1점|1개",
    r"품질\s*(나쁜|이\s*나쁜)",
    r"잘\s*안\s*돼|안\s*되는|못\s*쓰겠|못써|쓸모없어",
    r"수량\s*적음|색이\s*다름|색상\s*상이|성능\s*과장|원산지\s*불명확",
    r"색감\s*차이|과포장|과점착|색번짐|이염",
    r"후회|실패|구매\s*실패|돈\s*낭비|불합리",
    r"사지\s*마|추천\s*안",
    r"뒤틀|휘어",
    r"금방\s*(망가|부서)|오래\s*못\s*가",
    r"변질|변질되|변질됐|부서지|부서졌",
    r"녹았|녹이\s*슬",
    r"생각보다\s*별로|기대\s*이하|글쎄|그냥\s*저냥",
    r"이상해|별로야|별로네",
]

# ── 긍정 키워드 ──────────────────────────────────────────────
POSITIVE_PATTERNS = [
    r"좋아요|좋았|만족|추천|재구매|최고|훌륭|완벽|편리|예뻐",
    r"괜찮(았|네|아요|습니다)?",
    r"가성비\s*(최고|좋|굿|짱|갓성비)?",
    r"품질\s*(좋|굿|최고|만족|좋아|짱)",
    r"탁월|우수|뛰어나|최상|압도적|놀랍|감탄",
    r"만족스럽|흡족|대만족|마음에\s*든|기대\s*이상",
    r"실망\s*없|후회\s*없|잘\s*샀|강력\s*추천|재구매\s*의사",
    r"기쁘|즐겁|설렌|뿌듯|감동|감격|황홀|짜릿|신난",
    r"친절|상냥|센스\s*있|배려|정성|세심|꼼꼼|믿음직|전문적",
    r"돈값|이득|효과\s*만점|효율적|경제적|합리적",
    r"아름답|세련|고급스럽|멋있|깔끔|심플|귀엽|화사|이쁘",
    r"편하|사용하기\s*쉽|직관적|간편|부드럽|가볍|착용감|실용적",
    r"강추|적극\s*추천|또\s*살|반드시\s*사야|선물하고\s*싶",
    r"굿|대박|짱|갓|레전드|찐|꿀템|득템|최애|맘에\s*(들|쏙)",
    r"AS가\s*완벽|응대가\s*빠르|처리가\s*빠르",
    r"기대보다\s*훨씬|사길\s*잘했|삶의\s*질이\s*올라",
    r"완전\s*좋|지림|지려|최적|최상의|행복|사랑",
]

# ── 홍보성 키워드 ──────────────────────────────────────────────
PROMO_PATTERNS = [
    r"제공.{0,5}받|협찬|체험단|서포터즈",
    r"소정의\s*원고료|원고료.*지급|광고.*포함",
    r"링크.*통해.*구매|할인\s*코드|쿠폰\s*코드",
    r"#\s*(ad|광고|협찬|제공|유료)",
    r"프로필\s*링크|바이오\s*링크|링크\s*걸어",
    r"DM\s*(주세요|문의|으로)|댓글로\s*문의",
    r"공\s*구|공동\s*구매|선착\s*순|한정\s*수량",
    r"내\s*돈\s*내\s*산|솔직\s*후기|진짜\s*후기|광고\s*아님|협찬\s*아님",
    r"오늘만\s*이\s*가격|지금만\s*할인|마감\s*임박|품절\s*임박|역대급\s*할인",
    r"후기\s*남겨|태그\s*해|팔로우\s*하면|좋아요\s*누르면|저장\s*해",
    r"부모님께\s*사드|선물했더니\s*좋아|온\s*가족이\s*쓰",
    r"홍보|리뷰어|내돈내산\s*아님|무료로\s*받",
    r"하울|추천템|인기템|꿀템\s*추천|베스트|신상품\s*추천",
    r"도전|챌린지|이벤트|세일|프로모션|특가",
    r"매장\s*(옆|근처|앞|뒤|주변|위치|주차|방문|영업시간)",
    r"이웃추가|서로이웃|구독하기|댓글\s*남겨주세요|공감\s*눌러",
    r"포스팅|블로그\s*포스트|오늘의\s*포스팅|리뷰\s*포스팅",
    r"원고료|고료|소정의",
    r"카페\s*회원|카페\s*가입|카페\s*링크|카페\s*공구",
    r"공구\s*오픈|공구\s*마감|공구\s*진행중|공구\s*참여",
    r"네이버\s*쇼핑|스마트스토어|스토어팜",
    r"최저가\s*확인|가격\s*비교|쇼핑\s*검색",
    r"네이버\s*체험단|레뷰|강남언니|위블|미블",
    r"서포터즈\s*활동|서포터즈\s*선정|앰배서더",
    r"상단\s*링크|링크\s*참고|아래\s*링크|위\s*링크",
    r"자세한\s*내용은|더\s*보러가기|전체\s*후기는",
    r"이상으로.*리뷰|이상.*포스팅|이상.*후기",
    r"도움이\s*됐으면|도움이\s*되셨으면|참고가\s*됐으면",
    r"읽어주셔서\s*감사|방문해\s*주셔서\s*감사",
]

TITLE_PROMO_KW = ["추천", "하울", "꿀템", "인생템", "갓성비", "득템", "베스트", "추천템"]

# 구글시트 키워드 (단순 문자열) + 골드셋 키워드
_SHEET_NEG_KW = _sheet_kw.get("neg", [])
_SHEET_POS_KW = _sheet_kw.get("pos", [])
PROMO_KW    = list(set(_sheet_kw.get("promo", [])))
SHEET_EXCLUDE_KW = _sheet_kw.get("exclude", [])

_goldset_kw = extract_goldset_keywords()
_GOLDSET_NEG_KW = _goldset_kw.get("부정", [])
_GOLDSET_POS_KW = _goldset_kw.get("긍정", [])

# 통합 매칭 함수
def count_negative(text, exclude_words=None):
    exclude = exclude_words or []
    cnt = 0
    for p in NEGATIVE_PATTERNS:
        m = re.search(p, text)
        if m and m.group() not in exclude:
            cnt += 1
    for kw in _SHEET_NEG_KW + _GOLDSET_NEG_KW:
        if kw in text and kw not in exclude:
            cnt += 1
    return cnt

def count_positive(text, exclude_words=None):
    exclude = exclude_words or []
    cnt = 0
    for p in POSITIVE_PATTERNS:
        m = re.search(p, text)
        if m and m.group() not in exclude:
            cnt += 1
    for kw in _SHEET_POS_KW + _GOLDSET_POS_KW:
        if kw in text and kw not in exclude:
            cnt += 1
    return cnt

def has_negative(text):
    return count_negative(text) > 0

def has_positive(text):
    return count_positive(text) > 0

def is_promotional(item: dict) -> bool:
    title = clean_text(item.get("title", ""))
    desc  = clean_text(item.get("description", ""))
    full  = title + " " + desc
    promo_hit = sum(1 for kw in PROMO_KW if kw in full)
    pattern_hit = sum(1 for p in PROMO_PATTERNS if re.search(p, full))
    neg_hit = count_negative(full)
    title_promo = sum(1 for kw in TITLE_PROMO_KW if kw in title)
    if (promo_hit >= 1 or pattern_hit >= 1 or title_promo >= 1) and neg_hit <= 1:
        return True
    return False


LABEL_MAP = {
    "positive":"긍정","pos":"긍정","LABEL_2":"긍정","긍정":"긍정",
    "negative":"부정","neg":"부정","LABEL_0":"부정","부정":"부정",
    "neutral":"중립","neu":"중립","LABEL_1":"중립","중립":"중립",
    "부정":"부정","긍정":"긍정",
}

def rule_based(text: str, exclude_words=None):
    neg = count_negative(text, exclude_words)
    pos = count_positive(text, exclude_words)
    # 다이소 맥락 부정 가중
    daiso_in_text = any(v in text for v in DAISO_VARIANTS)
    if daiso_in_text and neg > 0:
        neg += 1
    if neg > pos:  return "부정", min(0.65 + neg * 0.08, 0.98)
    if pos > neg:  return "긍정", min(0.60 + pos * 0.08, 0.98)
    return "중립", 0.50

def get_reason_sentence(full_text: str, sentiment: str) -> str:
    """감성 판단 근거가 되는 문장 한 줄 추출."""
    patterns = NEGATIVE_PATTERNS if sentiment == "부정" else POSITIVE_PATTERNS if sentiment == "긍정" else []
    if not patterns:
        return ""
    sentences = re.split(r'[.!?\n]+', full_text)
    best_sent, best_cnt = "", 0
    for s in sentences:
        s = s.strip()
        if not s:
            continue
        if sentiment == "부정" and "다이소" not in full_text and "다이소" not in s:
            continue
        cnt = sum(1 for p in patterns if re.search(p, s))
        if cnt > best_cnt:
            best_cnt = cnt
            best_sent = s
    if best_sent and len(best_sent) > 80:
        best_sent = best_sent[:80] + "…"
    return best_sent

def ensemble_sentiment(roberta_output, full_text: str, threshold: int, exclude_words=None) -> tuple:
    votes = {"긍정": 0.0, "부정": 0.0, "중립": 0.0}
    exclude = exclude_words or []

    roberta_neg_prob = 0.0
    if roberta_output:
        try:
            for it in roberta_output:
                lbl = LABEL_MAP.get(it["label"])
                if lbl:
                    votes[lbl] += it["score"] * 1.8
                    if lbl == "부정":
                        roberta_neg_prob = it["score"]
        except Exception:
            pass

    rule_lbl, rule_sc = rule_based(full_text, exclude_words=exclude)
    votes[rule_lbl] += rule_sc * 1.5

    total = sum(votes.values())
    if total == 0:
        return "중립", 50, ""
    best  = max(votes, key=votes.get)
    score = round(votes[best] / total * 100)
    neg_kw_cnt = count_negative(full_text, exclude)

    if neg_kw_cnt >= 3:
        return "부정", max(score, 75), get_reason_sentence(full_text, "부정")
    if roberta_neg_prob >= 0.5 and neg_kw_cnt >= 1:
        return "부정", max(score, 70), get_reason_sentence(full_text, "부정")
    if roberta_neg_prob >= 0.4 and neg_kw_cnt >= 2:
        return "부정", max(score, 65), get_reason_sentence(full_text, "부정")

    if score < threshold and best == "부정":
        return "중립", max(score - 10, 40), ""

    reason = get_reason_sentence(full_text, best)
    return best, score, reason


# ============================
# 블로그 본문 크롤링 (합법적 범위)
# ============================
def fetch_blog_body(url: str) -> str:
    """네이버 블로그 본문 텍스트 추출. 실패 시 빈 문자열 반환."""
    try:
        # 네이버 블로그 모바일 버전 (iframe 없이 본문 접근 가능)
        if "blog.naver.com" in url:
            parts = url.replace("https://", "").replace("http://", "").split("/")
            if len(parts) >= 3:
                blog_id = parts[1]
                log_no = parts[2].split("?")[0]
                mobile_url = f"https://m.blog.naver.com/{blog_id}/{log_no}"
            else:
                return ""
        else:
            return ""
        headers = {
            "User-Agent": "Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36"
        }
        resp = requests.get(mobile_url, headers=headers, timeout=8)
        if resp.status_code != 200:
            return ""
        # 본문 영역 추출
        text = resp.text
        # se-main-container 또는 post-view 영역
        body_match = re.search(r'<div class="se-main-container">(.*?)</div>\s*</div>\s*</div>', text, re.DOTALL)
        if not body_match:
            body_match = re.search(r'<div class="post_ct"[^>]*>(.*?)</div>', text, re.DOTALL)
        if not body_match:
            body_match = re.search(r'<div class="se_component_wrap">(.*?)</div>\s*</div>', text, re.DOTALL)
        if body_match:
            body_html = body_match.group(1)
        else:
            # 전체에서 텍스트 추출 (fallback)
            body_html = text
        # HTML 태그 제거
        body_text = re.sub(r'<[^>]+>', ' ', body_html)
        body_text = re.sub(r'&[a-zA-Z]+;', ' ', body_text)
        body_text = re.sub(r'\s+', ' ', body_text).strip()
        return body_text[:2000]  # 최대 2000자
    except Exception:
        return ""


# ============================
# 다이소 관련성 필터
# ============================
DAISO_VARIANTS = ["다이소", "DAISO", "daiso"]

# 다이소 상품/불만 맥락 판별용 키워드
DAISO_CONTEXT_KW = [
    # 상품/구매 맥락
    "제품","상품","구매","구입","샀","사용","써봤","써보","품질","가격","원짜리",
    "천원","이천원","삼천원","오천원","만원","리뷰","후기","언박싱","하울",
    "개봉","포장","디자인","색상","사이즈","크기","재질","내구성","마감",
    "불량","고장","교환","환불","깨졌","망가","부러","터졌","좋아","만족",
    "추천","최고","괜찮","예쁘","편리","가성비","별로","실망","후회",
    "수납","정리","주방","욕실","문구","인테리어","공구","전자","화장품",
    "식품","생활","세제","청소","빨래","건조","조명","충전","케이블",
    "보관","용기","그릇","컵","접시","냄비","팬","칼","가위","테이프",
    "후크","선반","바구니","서랍","파일","노트","펜","스티커","봉투",
]

# 다이소가 장소/위치로만 언급된 경우 제외
DAISO_LOCATION_EXCLUDE = [
    "다이소 옆","다이소 근처","다이소 앞","다이소 뒤","다이소 건너편",
    "다이소 맞은편","다이소 위","다이소 아래","다이소 건물","다이소 빌딩",
    "다이소 골목","다이소 사거리","다이소 교차로","다이소 버스",
    "다이소 지하철","다이소 역","다이소 정류장",
    "다이소 점포","다이소 매장","다이소 오픈","다이소 영업","다이소 위치",
    "다이소 주차","다이소 방문","다이소 알바","다이소 직원","다이소 채용",
    "다이소 1층", "다이소 2층", "다이소 3층", "다이소 4층", "다이소 지하",
]

def is_daiso_related(item: dict) -> bool:
    """관련성 점수가 임계값 이상이면 통과 (기존 이진 판정 → 점수제로 전환).
    실제 점수 계산은 daiso_relevance_score, 이 함수는 통과 여부만 반환.
    """
    score = daiso_relevance_score(item)
    item["_relevance"] = score  # 결과 카드 표시용
    return score >= DAISO_RELEVANCE_THRESHOLD

# 관련성 통과 임계값 (튜닝 지점)
DAISO_RELEVANCE_THRESHOLD = 5

# 다른 매장/브랜드 (다이소와 함께 언급되면 감점)
_OTHER_BRAND_KW = [
    "이케아", "IKEA", "무인양품", "무지", "MUJI", "코스트코", "COSTCO",
    "홈플러스", "이마트", "롯데마트", "쿠팡", "11번가", "지마켓", "옥션",
    "네이버쇼핑", "마트몰", "아트박스", "모닝글로리", "핫트랙스", "교보",
    "올리브영", "왓슨스", "롭스", "세븐일레븐", "GS25", "CU", "이마트24",
]

def _learned_exclude_kws_cached() -> list:
    """세션 내 캐시된 제외 학습 키워드 반환."""
    if "_learned_exclude_cache" in st.session_state:
        return st.session_state["_learned_exclude_cache"]
    try:
        kws = _extract_excluded_features(min_count=2, top_n=30)
    except Exception:
        kws = []
    st.session_state["_learned_exclude_cache"] = kws
    return kws

def daiso_relevance_score(item: dict) -> int:
    """다이소 관련성 점수 산정. 높을수록 유효 글일 가능성 큼.

    가산 요소:
    + 5: 제목에 다이소 있음
    + 5: 본문에 다이소 3회 이상
    + 3: 본문에 다이소 1~2회
    +10: 유효 품번 명시
    + 8: 품명DB 완전 매칭
    + 3×N: 다이소 근처(±40자)에 상품 맥락어 (최대 3개 카운트)
    + 2: 부정/긍정 패턴 hit
    감점 요소:
    -10: 장소 문맥에서만 등장
    - 3: 다른 브랜드가 함께 언급 (다이소 언급이 소수일 때)
    - 5: 제외 학습 특징어 1개 등장
    -10: 제외 학습 특징어 2개+ 등장
    """
    title = re.sub(r"<[^>]+>", "", item.get("title", "") or "")
    desc = re.sub(r"<[^>]+>", "", item.get("description", "") or "")
    raw = (title + " " + desc)
    raw_upper = raw.upper()

    score = 0

    # 다이소 언급 카운트
    daiso_count = 0
    for v in DAISO_VARIANTS:
        daiso_count += raw.count(v) if v != v.upper() else raw_upper.count(v.upper())
    # (한글/영문 중복 방지: 대략 최대치 사용)
    daiso_count = max(raw.count("다이소"), raw_upper.count("DAISO"))

    # 제목에 다이소
    title_upper = title.upper()
    if "다이소" in title or "DAISO" in title_upper:
        score += 5

    # 본문 다이소 언급
    if daiso_count >= 3:
        score += 5
    elif daiso_count >= 1:
        score += 3

    # 유효 품번 명시 (다이소 언급 없어도 강력한 시그널)
    if VALID_PRODUCT_CODES:
        codes = re.findall(r"\b(\d{6,8})\b", raw)
        for c in codes:
            if c in VALID_PRODUCT_CODES:
                score += 10
                break

    # 품명DB 매칭 (짧은 품명 오탐 방지 로직 재사용)
    if PRODUCT_NAME_INDEX:
        for pn in PRODUCT_NAME_INDEX:
            name = pn["품명"]
            if len(name) >= 4 and name in raw:
                score += 8
                break
            elif 2 <= len(name) <= 3 and _match_short_name(raw, name):
                score += 6
                break

    # 다이소 ±40자 이내의 상품 맥락어
    if daiso_count > 0:
        ctx_hits = 0
        for m in re.finditer(r"다이소|DAISO", raw, flags=re.IGNORECASE):
            near = raw[max(0, m.start() - 40): m.end() + 40]
            for kw in DAISO_CONTEXT_KW:
                if kw in near:
                    ctx_hits += 1
                    if ctx_hits >= 3:
                        break
            if ctx_hits >= 3:
                break
        score += min(ctx_hits, 3) * 3

    # 감성 패턴 hit (은근한 관련성 시그널)
    try:
        if has_negative(raw) or has_positive(raw):
            score += 2
    except Exception:
        pass

    # ── 감점 ──
    # 장소 문맥
    loc_hits = sum(1 for loc in DAISO_LOCATION_EXCLUDE if loc in raw)
    if loc_hits >= 1 and daiso_count <= 2:
        score -= 10

    # 다른 브랜드와 함께 언급 (다이소 언급이 소수일 때만 감점)
    other_brand_hits = sum(1 for b in _OTHER_BRAND_KW if b in raw or b.upper() in raw_upper)
    if other_brand_hits >= 1 and daiso_count <= 1:
        score -= 3
    if other_brand_hits >= 3 and daiso_count <= 2:
        score -= 3

    # 제외 학습 특징어
    try:
        learned = _learned_exclude_kws_cached()
        if learned:
            ex_hit = sum(1 for kw in learned if kw and kw in raw)
            if ex_hit >= 2:
                score -= 10
            elif ex_hit >= 1:
                score -= 5
    except Exception:
        pass

    return score

def build_naver_query(raw_keywords: str) -> str:
    kw = raw_keywords.strip()
    has_daiso = any(v in kw for v in DAISO_VARIANTS)
    if not has_daiso:
        kw = "다이소 " + kw
    return kw


# ============================
# 네이버 블로그 수집 (페이징)
# ============================
def collect_naver_paged(query: str, search_type: str, total: int) -> list:
    all_items = []
    per_page  = 100
    start_idx = 1
    label = "블로그"

    while len(all_items) < total:
        if start_idx > 1000:
            break
        remaining = total - len(all_items)
        fetch_cnt = min(per_page, remaining, 1000 - start_idx + 1)
        if fetch_cnt <= 0:
            break

        url     = f"https://openapi.naver.com/v1/search/{search_type}.json"
        headers = {"X-Naver-Client-Id": NAVER_CLIENT_ID, "X-Naver-Client-Secret": NAVER_CLIENT_SECRET}
        params  = {"query": query, "display": fetch_cnt, "start": start_idx, "sort": "date"}
        try:
            resp  = requests.get(url, headers=headers, params=params, timeout=10)
            items = resp.json().get("items", [])
        except Exception:
            break

        if not items:
            break

        for item in items:
            item["출처"]   = label
            item["검색어"] = query
        all_items.extend(items)
        start_idx += fetch_cnt

        if len(items) < fetch_cnt:
            break

    return all_items[:total]


# ============================================== 네이버 블로그 본문 자동 병합
@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_blog_body_cached(url: str) -> str:
    """캐시 래퍼 (URL당 1시간)."""
    return fetch_blog_body(url)

def enrich_blog_with_body(items: list, max_items: int = 50) -> list:
    """네이버 블로그 검색결과에 본문 텍스트를 병합.
    - blog.naver.com URL만 시도 (다른 도메인은 스킵)
    - description 뒤에 '[본문] ...' 형태로 추가 → 감성분석/필터/AI 근거 확보
    - 병렬로 처리해 속도 확보
    """
    if not items:
        return items
    limited = items[:max_items]
    urls = []
    idxs = []
    for i, it in enumerate(limited):
        link = it.get("link", "") or ""
        if "blog.naver.com" in link:
            urls.append(link)
            idxs.append(i)
    if not urls:
        return items
    bodies = {}
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=6) as ex:
            for u, body in zip(urls, ex.map(_fetch_blog_body_cached, urls)):
                bodies[u] = body or ""
    except Exception:
        for u in urls:
            bodies[u] = _fetch_blog_body_cached(u) or ""
    for i in idxs:
        u = limited[i].get("link", "")
        body = bodies.get(u, "")
        if body:
            base = limited[i].get("description", "") or ""
            limited[i]["blog_body"] = body
            limited[i]["description"] = (base + " [본문] " + body)[:3500]
    return items


# ============================
# 네이버 카페 수집 (페이징)
# ============================
def collect_cafe_paged(query: str, total: int) -> list:
    all_items = []
    per_page  = 100
    start_idx = 1

    while len(all_items) < total:
        if start_idx > 1000:
            break
        remaining = total - len(all_items)
        fetch_cnt = min(per_page, remaining, 1000 - start_idx + 1)
        if fetch_cnt <= 0:
            break

        url     = "https://openapi.naver.com/v1/search/cafearticle.json"
        headers = {"X-Naver-Client-Id": NAVER_CLIENT_ID, "X-Naver-Client-Secret": NAVER_CLIENT_SECRET}
        params  = {"query": query, "display": fetch_cnt, "start": start_idx, "sort": "date"}
        try:
            resp  = requests.get(url, headers=headers, params=params, timeout=10)
            items = resp.json().get("items", [])
        except Exception:
            break

        if not items:
            break

        for item in items:
            item["출처"]   = "카페"
            item["검색어"] = query
            item["channel"] = item.get("cafename", "")

        all_items.extend(items)
        start_idx += fetch_cnt

        if len(items) < fetch_cnt:
            break

    return all_items[:total]


# ============================
# YouTube
# ============================
def search_youtube(query: str, max_results: int = 30) -> list:
    if not YOUTUBE_API_KEY: return []
    all_items = []
    page_token = None
    while len(all_items) < max_results:
        params = {
            "key": YOUTUBE_API_KEY, "q": query, "part": "snippet",
            "type": "video", "maxResults": min(50, max_results - len(all_items)),
            "order": "date", "relevanceLanguage": "ko", "regionCode": "KR"
        }
        if page_token:
            params["pageToken"] = page_token
        try:
            resp = requests.get("https://www.googleapis.com/youtube/v3/search", params=params, timeout=10)
            data = resp.json()
        except Exception:
            break
        if "error" in data:
            break
        items = data.get("items", [])
        if not items:
            break
        all_items.extend(items)
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    all_items = all_items[:max_results]
    video_ids = [i["id"]["videoId"] for i in all_items if i.get("id", {}).get("videoId")]
    stats_map = {}
    if video_ids:
        for i in range(0, len(video_ids), 50):
            chunk = video_ids[i:i+50]
            try:
                for sv in requests.get("https://www.googleapis.com/youtube/v3/videos", params={
                    "key": YOUTUBE_API_KEY, "id": ",".join(chunk), "part": "statistics"
                }, timeout=10).json().get("items", []):
                    stats_map[sv["id"]] = sv.get("statistics", {})
            except Exception:
                pass
    results = []
    for item in all_items:
        vid_id  = item.get("id", {}).get("videoId", "")
        snippet = item.get("snippet", {})
        stats   = stats_map.get(vid_id, {})
        pub_raw = snippet.get("publishedAt", "")
        try:   pub_dt = datetime.strptime(pub_raw[:10], "%Y-%m-%d"); pub_str = pub_dt.strftime("%Y-%m-%d")
        except: pub_dt = None; pub_str = pub_raw[:10]
        results.append({
            "출처":"유튜브","검색어":query,"video_id":vid_id,
            "title":snippet.get("title",""),
            "description":snippet.get("description","")[:300],
            "channel":snippet.get("channelTitle",""),
            "thumbnail":snippet.get("thumbnails",{}).get("medium",{}).get("url",""),
            "link":f"https://www.youtube.com/watch?v={vid_id}",
            "날짜":pub_str,"pub_dt":pub_dt,
            "views":int(stats.get("viewCount",0) or 0),
            "likes":int(stats.get("likeCount",0) or 0),
            "comments":int(stats.get("commentCount",0) or 0),
        })
    return results


# ============================================== 유튜브 자막 수집 (youtube-transcript-api)
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_youtube_transcript(video_id: str, max_chars: int = 3000) -> str:
    """유튜브 자막을 텍스트로 반환. 한국어 우선, 자동생성 자막도 허용.
    자막이 없거나 실패하면 빈 문자열 반환.
    """
    if not video_id:
        return ""
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
    except ImportError:
        return ""
    try:
        # 한국어 우선, 없으면 영어, 그것도 없으면 아무거나(자동생성 포함)
        try:
            transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
            transcript = None
            for lang in ["ko", "ko-KR"]:
                try:
                    transcript = transcript_list.find_transcript([lang])
                    break
                except Exception:
                    continue
            if transcript is None:
                # 자동생성 자막 포함
                try:
                    transcript = transcript_list.find_generated_transcript(["ko"])
                except Exception:
                    transcript = None
            if transcript is None:
                # 마지막 폴백: 아무거나
                for t in transcript_list:
                    transcript = t
                    break
            if transcript is None:
                return ""
            entries = transcript.fetch()
        except Exception:
            # list_transcripts 실패 시 직접 시도
            entries = YouTubeTranscriptApi.get_transcript(video_id, languages=["ko", "ko-KR", "en"])
        text = " ".join(seg.get("text", "").strip() for seg in entries)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:max_chars]
    except Exception:
        return ""

def enrich_youtube_with_transcripts(items: list, max_items: int = 50) -> list:
    """유튜브 검색 결과에 자막 텍스트(description에 병합)를 추가.
    자막이 있으면 description 뒤에 '[자막] ...' 형태로 붙여 감성분석 근거를 강화.
    max_items: 자막을 시도할 최대 개수 (과도한 API 호출 방지)
    """
    if not items:
        return items
    limited = items[:max_items]
    for it in limited:
        vid = it.get("video_id", "")
        if not vid:
            continue
        transcript = fetch_youtube_transcript(vid, max_chars=2000)
        if transcript:
            base_desc = it.get("description", "") or ""
            it["transcript"] = transcript
            # 감성분석/필터가 description을 보므로 병합해서 근거 확보
            it["description"] = (base_desc + " [자막] " + transcript)[:3500]
    return items


# ============================================== 유튜브 댓글 수집 (commentThreads)
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_youtube_comments(video_id: str, max_comments: int = 30) -> list:
    """유튜브 영상의 상위 댓글을 반환. 광고 없이 진짜 소비자 목소리 소스.
    반환: [{"text","author","like_count","publishedAt"}, ...]
    """
    if not YOUTUBE_API_KEY or not video_id:
        return []
    try:
        resp = requests.get(
            "https://www.googleapis.com/youtube/v3/commentThreads",
            params={
                "key": YOUTUBE_API_KEY,
                "part": "snippet",
                "videoId": video_id,
                "maxResults": min(max_comments, 100),
                "order": "relevance",
                "textFormat": "plainText",
            },
            timeout=10,
        )
        data = resp.json()
    except Exception:
        return []
    if "error" in data:
        return []
    out = []
    for it in data.get("items", []):
        try:
            top = it["snippet"]["topLevelComment"]["snippet"]
            out.append({
                "text": top.get("textDisplay") or top.get("textOriginal") or "",
                "author": top.get("authorDisplayName", ""),
                "like_count": int(top.get("likeCount", 0) or 0),
                "publishedAt": top.get("publishedAt", "")[:10],
            })
        except Exception:
            continue
    return out

def comments_to_items(video_meta: dict, comments: list) -> list:
    """댓글 리스트를 크롤링 아이템 스키마로 변환.
    영상 1개당 댓글 N개 = 아이템 N개.
    - 상품 매칭은 영상 제목/자막을 컨텍스트로 이어붙임 (댓글 자체는 짧음)
    - link는 영상 URL + 댓글 앵커
    """
    items = []
    vid = video_meta.get("video_id", "")
    v_link = f"https://www.youtube.com/watch?v={vid}"
    v_title = video_meta.get("title", "")
    v_desc = video_meta.get("description", "")[:800]  # 자막 병합된 상태
    for i, c in enumerate(comments):
        text = c.get("text", "").strip()
        if not text or len(text) < 5:
            continue
        # 댓글 + 영상 컨텍스트 = 감성/상품 분석 근거
        merged_desc = f"[댓글] {text}\n[영상] {v_title} {v_desc}"[:3500]
        items.append({
            "출처": "유튜브 댓글",
            "검색어": video_meta.get("검색어", ""),
            "video_id": vid,
            "title": text[:80] + ("…" if len(text) > 80 else ""),
            "description": merged_desc,
            "channel": c.get("author", ""),
            "thumbnail": video_meta.get("thumbnail", ""),
            "link": f"{v_link}&lc={i}",  # 유사 앵커 (실제 comment id는 별도 필요)
            "날짜": c.get("publishedAt", "") or video_meta.get("날짜", ""),
            "views": 0,
            "likes": c.get("like_count", 0),
            "comments": 0,
        })
    return items


# ============================================== 네이버 지식iN 검색
def collect_kin_paged(query: str, total: int) -> list:
    """네이버 검색 API 지식iN 탭.
    ※ 네이버 API는 지식iN에 sort=date를 지원하지 않아 관련도 순만 가능.
    """
    all_items = []
    per_page = 100
    start_idx = 1
    label = "지식iN"
    while len(all_items) < total:
        if start_idx > 1000:
            break
        remaining = total - len(all_items)
        fetch_cnt = min(per_page, remaining, 1000 - start_idx + 1)
        if fetch_cnt <= 0:
            break
        url = "https://openapi.naver.com/v1/search/kin.json"
        headers = {"X-Naver-Client-Id": NAVER_CLIENT_ID, "X-Naver-Client-Secret": NAVER_CLIENT_SECRET}
        params = {"query": query, "display": fetch_cnt, "start": start_idx, "sort": "sim"}  # sim = 관련도
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=10)
            items = resp.json().get("items", [])
        except Exception:
            break
        if not items:
            break
        for item in items:
            item["출처"] = label
            item["검색어"] = query
        all_items.extend(items)
        start_idx += fetch_cnt
        if len(items) < fetch_cnt:
            break
    return all_items[:total]


# ============================================== 뽐뿌 자유게시판 검색 크롤러
_PPOMPPU_SEARCH_URL = "https://www.ppomppu.co.kr/search_bbs.php"

@st.cache_data(ttl=1800, show_spinner=False)
def collect_ppomppu_freeboard(query: str, total: int = 200) -> list:
    """뽐뿌 자유게시판(freeboard) 제목+본문 검색.
    - 최대 200개, 최신순
    - 게시글 상세 페이지 본문까지 병합해서 감성/상품 분석 근거 확보
    - 결과의 출처 필드는 '커뮤니티'로 라벨링 (UI 통일)
    """
    if not query:
        return []
    results = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121 Safari/537.36",
        "Referer": "https://www.ppomppu.co.kr/",
    }
    per_page = 20
    total = min(total, 500)
    max_pages = (total + per_page - 1) // per_page
    # 검색 리스트 페이지 순회
    for page in range(1, max_pages + 1):
        if len(results) >= total:
            break
        try:
            resp = requests.get(
                _PPOMPPU_SEARCH_URL,
                params={
                    "search_type": "sub_memo",   # 제목+본문
                    "bbs_id": "freeboard",
                    "keyword": query,
                    "page_number": page,
                    "order_type": "date",
                },
                headers=headers, timeout=10,
            )
            if resp.status_code != 200:
                break
            html = resp.text
        except Exception:
            break
        page_items = _parse_ppomppu_search_list(html)
        if not page_items:
            break
        for it in page_items:
            if len(results) >= total:
                break
            it["출처"] = "커뮤니티"
            it["검색어"] = query
            # 본문 병합 (병렬로 처리)
            results.append(it)
        time.sleep(0.4)   # rate limit 예방
    # 상세 본문 병렬 수집
    if results:
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=6) as ex:
                bodies = list(ex.map(lambda u: _fetch_ppomppu_body(u), [r["link"] for r in results]))
            for r, body in zip(results, bodies):
                if body:
                    r["description"] = (r.get("description", "") + " [본문] " + body)[:3500]
        except Exception:
            pass
    return results

def _parse_ppomppu_search_list(html: str) -> list:
    """뽐뿌 검색 결과 리스트 파싱. HTML 구조 변경 시 이 함수만 손보면 됨."""
    items = []
    # 게시글 링크: /zboard/view.php?id=freeboard&no=NNN
    # 리스트 항목은 각 <tr>로 렌더링됨
    for m in re.finditer(
        r'<a[^>]+href="([^"]*view\.php\?id=freeboard[^"]*)"[^>]*>(.*?)</a>',
        html, re.DOTALL | re.IGNORECASE
    ):
        href = m.group(1)
        title_html = m.group(2)
        # HTML 태그 제거
        title = re.sub(r"<[^>]+>", "", title_html)
        title = re.sub(r"\s+", " ", title).strip()
        if not title or len(title) < 3:
            continue
        # 절대 URL로
        if href.startswith("/"):
            href = "https://www.ppomppu.co.kr" + href
        elif not href.startswith("http"):
            href = "https://www.ppomppu.co.kr/zboard/" + href.lstrip("./")
        # 중복 제거
        if any(it["link"] == href for it in items):
            continue
        items.append({
            "title": title,
            "link": href,
            "description": "",
            "날짜": "",  # 상세 페이지에서 채움
        })
    return items

@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_ppomppu_body(url: str) -> str:
    """뽐뿌 게시글 상세 페이지에서 본문 텍스트 추출. 실패 시 빈 문자열."""
    if not url or "ppomppu.co.kr" not in url:
        return ""
    try:
        resp = requests.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121 Safari/537.36",
            "Referer": "https://www.ppomppu.co.kr/",
        }, timeout=8)
        if resp.status_code != 200:
            return ""
        html = resp.text
        # 본문 영역 대표 후보들 (뽐뿌 HTML 변경 대비 여러 패턴)
        patterns = [
            r'<td[^>]+class="board-contents"[^>]*>(.*?)</td>',
            r'<div[^>]+class="board-contents"[^>]*>(.*?)</div>',
            r'<td[^>]*id="quote"[^>]*>(.*?)</td>',
            r'<div[^>]+id="quote"[^>]*>(.*?)</div>',
        ]
        body = ""
        for p in patterns:
            m = re.search(p, html, re.DOTALL | re.IGNORECASE)
            if m:
                body = m.group(1)
                break
        if not body:
            return ""
        body = re.sub(r"<[^>]+>", " ", body)
        body = re.sub(r"&[a-zA-Z]+;", " ", body)
        body = re.sub(r"\s+", " ", body).strip()
        return body[:2000]
    except Exception:
        return ""


def _fetch(task, display_count):
    tp, kw, label = task
    if tp == "blog":
        blog_items = collect_naver_paged(kw, "blog", display_count)
        # 블로그 본문 자동 병합 → description에 [본문] ... 붙여 감성/필터/AI 근거 강화
        try:
            blog_items = enrich_blog_with_body(blog_items, max_items=min(30, len(blog_items)))
        except Exception:
            pass
        return label, kw, blog_items
    if tp == "cafe": return label, kw, collect_cafe_paged(kw, display_count)
    if tp == "kin":  return label, kw, collect_kin_paged(kw, display_count)
    if tp == "ppomppu":
        return label, kw, collect_ppomppu_freeboard(kw, min(display_count, 200))
    if tp == "yt":
        yt_items = search_youtube(kw, max_results=min(display_count, 50))
        # 자막 자동 수집 → description에 병합 (감성분석/AI 근거 강화)
        try:
            yt_items = enrich_youtube_with_transcripts(yt_items, max_items=min(30, len(yt_items)))
        except Exception:
            pass
        return label, kw, yt_items
    if tp == "yt_comments":
        # 유튜브 영상 검색 후 각 영상의 상위 댓글을 수집해 아이템화
        yt_items = search_youtube(kw, max_results=min(display_count, 25))
        try:
            yt_items = enrich_youtube_with_transcripts(yt_items, max_items=min(15, len(yt_items)))
        except Exception:
            pass
        comments_items = []
        for v in yt_items:
            vid = v.get("video_id", "")
            if not vid:
                continue
            cs = fetch_youtube_comments(vid, max_comments=20)
            if cs:
                comments_items.extend(comments_to_items(v, cs))
            if len(comments_items) >= display_count:
                break
        return label, kw, comments_items[:display_count]
    return label, kw, []


# ============================
# 날짜 파싱 & 필터
# ============================
def parse_date(item: dict):
    ds = item.get("postdate") or item.get("pubDate", "")
    try:
        if len(ds) == 8:
            return datetime.strptime(ds, "%Y%m%d")
        return parsedate_to_datetime(ds).replace(tzinfo=None)
    except:
        return None

def filter_by_date(items: list, start_dt: date, end_dt: date) -> list:
    s = datetime(start_dt.year, start_dt.month, start_dt.day)
    e = datetime(end_dt.year,   end_dt.month,   end_dt.day, 23, 59, 59)
    result = []
    for item in items:
        if item.get("출처") == "카페":
            result.append(item)
            continue
        dt = item.get("pub_dt") if item.get("출처") == "유튜브" else parse_date(item)
        if dt and s <= dt <= e: result.append(item)
    return result

def clean_text(text: str) -> str:
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'&[a-zA-Z]+;', ' ', text)
    return text.strip()

def is_admin_excluded(item):
    url = item.get("link", "")
    if url in EXCLUDED_URLS_FROM_SHEET:
        return True
    full = clean_text(item.get("title","")) + " " + clean_text(item.get("description",""))
    all_exclude_kws = st.session_state.get("admin_exclude_kws", []) + SHEET_EXCLUDE_KW
    return any(kw in full for kw in all_exclude_kws)


# ============================
# 유심 관련 제외 필터
# ============================
USIM_EXCLUDE_KW = [
    "유심","USIM","유심칩","유심카드","심카드","SIM카드",
    "통신사","SKT","KT","LGU+","알뜰폰","eSIM","이심",
    "매장 옆","매장앞","매장 앞","매장 옆","옆 매장","옆가게","옆 매장",
    "유심기변","유심 기변","유심교체","유심 교체","유심 변경","유심변경",
    "해외유심","해외 유심","로밍유심","로밍 유심","글로벌유심","글로벌 유심",
    "다이소유심","다이소 유심","다이소심카드","다이소 심카드","정액","정력","이재명","대통령","주식","창업","소자본","부업","투잡","재테크","용돈벌이",
    "신용불량","사기","피해","보이스피싱","스미싱","금융사기","대출사기","투자사기","싱크대막힘","수리금액","변기관통기다이소","임신","개통가능한","선불폰",
    "균열","코킹"
]

def is_usim_related(item):
    text = (clean_text(item.get("title","")) + " " + clean_text(item.get("description",""))).upper()
    return any(kw.upper() in text for kw in USIM_EXCLUDE_KW)

# ============================
# 품번 추출
# ============================
DATE_PATS = [
    r'\b20\d{6}\b', r'\b\d{4}[-./]\d{2}[-./]\d{2}\b',
    r'\b\d{1,2}[-./]\d{1,2}[-./]\d{2,4}\b',
    r'\b\d{4}년\s*\d{1,2}월', r'\b\d{1,2}월\s*\d{1,2}일',
]
def is_date_like(t):
    for p in DATE_PATS:
        if re.fullmatch(p, t.strip()): return True
    return bool(re.fullmatch(r'20\d{6}', t.strip()))

YEAR_LIKE = {str(y) for y in range(2020, 2030)}

def extract_product_code(text):
    raw_nums = re.findall(r'\b(\d{4,11})\b', text)
    codes = []
    for c in raw_nums:
        if is_date_like(c):
            continue
        if c in YEAR_LIKE:
            continue
        if VALID_PRODUCT_CODES and c in VALID_PRODUCT_CODES:
            codes.append(c)
        elif not VALID_PRODUCT_CODES:
            codes.append(c)
    return ", ".join(dict.fromkeys(codes)) if codes else ""

def extract_price(text):
    prices = re.findall(r'\d{1,3}(?:,\d{3})*원', text)
    return ", ".join(dict.fromkeys(prices)) if prices else ""

SYNONYM_MAP = {
    "꽂이":"홀더","홀더":"꽂이","수납":"정리","정리":"수납",
    "바구니":"수납함","수납함":"바구니","케이스":"커버","커버":"케이스",
    "그릇":"용기","용기":"그릇","팬":"후라이팬","후라이팬":"팬",
    "집게":"클립","클립":"집게","수건":"타월","타월":"수건",
}

def extract_subcategory(text):
    if not SUBCATEGORIES: return ""
    found = [s for s in SUBCATEGORIES if s in text]
    if found: return ", ".join(dict.fromkeys(found))
    text_syn = text
    for w, s in SYNONYM_MAP.items(): text_syn = text_syn.replace(w, s)
    found2 = [s for s in SUBCATEGORIES if s in text_syn]
    if found2: return ", ".join(dict.fromkeys(found2))
    tokens = re.findall(r'[가-힣]{2,}', text)
    found3 = [s for s in SUBCATEGORIES if any(t in tokens for t in re.findall(r'[가-힣]{2,}', s) if len(t) >= 2)]
    if found3:
        found3.sort(key=lambda s: sum(1 for t in re.findall(r'[가-힣]{2,}', s) if t in tokens), reverse=True)
        return found3[0]
    return ""

def match_product_name(code):
    if PRODUCT_DB.empty or not code: return ""
    for c in [c.strip() for c in code.split(",")]:
        row = PRODUCT_DB[PRODUCT_DB["품번"].astype(str).str.strip() == c]
        if not row.empty: return row.iloc[0]["품명"]
    return ""


# ============================
# 품명 역매칭 (텍스트에서 품명 → 품번/소분류 추출)
# ============================
@st.cache_data(ttl=3600)
def _build_product_name_index():
    """품명 DB에서 2글자 이상 품명을 길이 내림차순으로 정렬한 리스트 생성."""
    if PRODUCT_DB.empty or "품명" not in PRODUCT_DB.columns:
        return []
    rows = []
    for _, r in PRODUCT_DB.iterrows():
        name = str(r.get("품명", "")).strip()
        if len(name) >= 2:
            rows.append({
                "품명": name,
                "품번": str(r.get("품번", "")).strip(),
                "소분류": str(r.get("소분류", "")).strip(),
            })
    # 긴 품명부터 매칭 (더 구체적인 것 우선)
    rows.sort(key=lambda x: len(x["품명"]), reverse=True)
    return rows

PRODUCT_NAME_INDEX = _build_product_name_index()

def match_by_product_name(text: str) -> dict:
    """텍스트에서 품명 DB의 품명이 포함되어 있으면 품번/품명/소분류 반환.
    (첫 매칭 하나만 반환 — 기존 호환용, 새 코드는 match_all_products 권장)
    """
    if not PRODUCT_NAME_INDEX:
        return {"품번": "", "품명": "", "소분류": ""}
    for item in PRODUCT_NAME_INDEX:
        pname = item["품명"]
        if len(pname) <= 3:
            # 짧은 품명은 단어 경계 체크 (조사·한자·공백 앞뒤에 오는 경우만)
            if _match_short_name(text, pname):
                return {"품번": item["품번"], "품명": pname, "소분류": item["소분류"]}
        elif pname in text:
            return {"품번": item["품번"], "품명": pname, "소분류": item["소분류"]}
    return {"품번": "", "품명": "", "소분류": ""}

def _match_short_name(text: str, name: str) -> bool:
    """2~3글자 짧은 품명의 오탐 방지 매칭.
    앞뒤가 한글이 아닌 경계(공백/문장부호/문자열끝/조사 앞)에서만 매칭.
    """
    if not name or name not in text:
        return False
    # 앞: 문장 시작 또는 비한글, 뒤: 조사/공백/문장부호/문자열끝
    # 짧은 명사는 조사가 붙는 게 자연스러움 (은/는/이/가/을/를/에/도/만/의)
    pattern = rf'(^|[^\w가-힣]){re.escape(name)}(은|는|이|가|을|를|에|도|만|의|과|와|로|으로|랑|이랑|이나|나|한|용|짜리|하나|두개|[\s\.,!?\)\]\}}"\'\-·]|$)'
    return re.search(pattern, text) is not None

def match_all_products(text: str, max_hits: int = 5) -> list:
    """텍스트에서 등장하는 모든 상품을 반환 (긴 품명부터, 중복 제거).
    반환: [{"품번","품명","소분류"}, ...] (최대 max_hits개)
    """
    if not PRODUCT_NAME_INDEX or not text:
        return []
    hits = []
    seen_codes = set()
    consumed_spans = []  # 이미 매칭한 부분은 다시 안 잡히도록 (짧은 품명이 긴 품명 안에 겹치는 경우 방지)
    for item in PRODUCT_NAME_INDEX:
        if len(hits) >= max_hits:
            break
        pname = item["품명"]
        code = item.get("품번", "")
        if code and code in seen_codes:
            continue
        # 짧은 품명은 경계 매칭
        if len(pname) <= 3:
            if not _match_short_name(text, pname):
                continue
            m = re.search(re.escape(pname), text)
            if not m:
                continue
            span = m.span()
        else:
            idx = text.find(pname)
            if idx < 0:
                continue
            span = (idx, idx + len(pname))
        # 이미 긴 품명에 포함된 위치라면 스킵
        if any(s[0] <= span[0] and span[1] <= s[1] for s in consumed_spans):
            continue
        consumed_spans.append(span)
        if code:
            seen_codes.add(code)
        hits.append({
            "품번": code,
            "품명": pname,
            "소분류": item.get("소분류", ""),
        })
    return hits


# ============================
# 엑셀 생성
# ============================
def create_excel(data: list, start_dt: date, end_dt: date) -> io.BytesIO:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "DAISO SNS ISSUE FINDER"
    headers = ["출처","검색어","소분류","품번","품명","가격언급","제목","링크","날짜","감성","확신도(%)","채널/카페명","조회수","좋아요","댓글수"]
    ws.append(headers)
    hf   = openpyxl.styles.Font(bold=True, color="0066CC", name="Malgun Gothic")
    hfil = openpyxl.styles.PatternFill(start_color="E8F1FB", end_color="E8F1FB", fill_type="solid")
    hbrd = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thin", color="0066CC"))
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c); cell.font = hf; cell.fill = hfil; cell.border = hbrd
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    col_bg  = {"긍정":"E8F5EE","부정":"FDEEEE","중립":"FFFBE8"}
    col_txt = {"긍정":"16A34A","부정":"DC2626","중립":"CA8A04"}
    for ri, row in enumerate(data, 2):
        ws.append([row.get(k,"") for k in ["출처","검색어","소분류","품번","품명","가격언급","title","link","날짜","감성","확신도","channel","views","likes","comments"]])
        s = row.get("감성","")
        if s in col_bg:
            ws.cell(ri,10).fill = openpyxl.styles.PatternFill(start_color=col_bg[s], end_color=col_bg[s], fill_type="solid")
            ws.cell(ri,10).font = openpyxl.styles.Font(color=col_txt[s], bold=True, name="Malgun Gothic")
    for letter, width in zip("ABCDEFGHIJKLMNO", [8,20,15,15,20,12,45,50,12,8,10,20,10,10,10]):
        ws.column_dimensions[letter].width = width
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ============================
# 헬퍼
# ============================
SENT_BADGE = {"긍정":"badge-pos","부정":"badge-neg","중립":"badge-neu"}

def icon(label: str) -> str:
    return f'<span class="section-title-icon">{label}</span>'

def fmt_score(score) -> str:
    try:
        return f"{int(round(float(score)))}%"
    except:
        return f"{score}%"


# ============================================== 관리자 분석 대시보드
def _render_admin_analytics(current_results: list):
    """관리자 탭 하단에 배치되는 3개 섹션.
    (1) 상품×이슈 매트릭스 (goldset 기반)
    (2) 감성분석 정확도 (goldset vs 앙상블)
    (3) 제외 학습 특징어 → 룰베이스 후보 채택
    """
    try:
        goldset_rows = load_goldset_from_sheet()
    except Exception:
        goldset_rows = []

    # ────────────────────── (1) 상품×이슈 매트릭스
    st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">상품별 고객 이슈 <span style="color:#A0AEC0;font-weight:500;">· 검토 목록 기반</span></div>', unsafe_allow_html=True)

    # 품번→(품명,카테고리) 조인 맵
    code_meta = {}
    if not PRODUCT_DB.empty:
        for _, row in PRODUCT_DB.iterrows():
            code = str(row.get("품번", "")).strip()
            if code:
                code_meta[code] = {
                    "품명": str(row.get("품명", "")).strip(),
                    "카테고리": str(row.get("소분류", "")).strip(),
                }

    # 상품별 이슈 집계
    prod_issues = {}   # code -> {label: Counter(issue)}
    for r in goldset_rows:
        code = str(r.get("product_code") or "").strip()
        label = (r.get("label") or "").strip()
        issue = (r.get("issue") or "").strip()
        if not code or label not in ("긍정", "부정", "중립"):
            continue
        prod_issues.setdefault(code, {"긍정": Counter(), "부정": Counter(), "중립": Counter()})
        prod_issues[code][label][issue or "(미기재)"] += 1

    if not prod_issues:
        st.info("아직 검토 목록에 상품 정보가 없습니다. 결과 카드에서 [확인 · 저장]으로 상품 · 이슈를 저장해 주세요.")
    else:
        # 총 건수 기준 정렬
        rows_sorted = sorted(
            prod_issues.items(),
            key=lambda kv: sum(sum(c.values()) for c in kv[1].values()),
            reverse=True,
        )
        table_html = '<div class="card" style="padding:0.4rem 0.6rem;overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-size:0.82rem;"><thead><tr style="background:#F1F5F9;color:#0F1B2D;">'
        for h in ["품번", "품명", "카테고리", "긍정", "부정", "중립", "주요 이슈 TOP3"]:
            table_html += f'<th style="text-align:left;padding:0.5rem 0.65rem;font-weight:700;font-size:0.72rem;letter-spacing:0.03em;text-transform:uppercase;color:#475569;">{h}</th>'
        table_html += '</tr></thead><tbody>'
        for code, cats in rows_sorted[:30]:
            pos = sum(cats["긍정"].values())
            neg = sum(cats["부정"].values())
            neu = sum(cats["중립"].values())
            meta = code_meta.get(code, {})
            # 부정 이슈 TOP3
            top_issues = cats["부정"].most_common(3) or cats["긍정"].most_common(3)
            issues_html = " · ".join(
                f'<span style="color:#DC2626;font-weight:600;">{i}</span> <span style="color:#7A879E;">({c})</span>'
                for i, c in top_issues if i and i != "(미기재)"
            ) or "<span style='color:#A0AEC0;'>—</span>"
            table_html += (
                f'<tr style="border-top:1px solid #E1E7F0;">'
                f'<td style="padding:0.5rem 0.65rem;font-family:monospace;color:#475569;">{code}</td>'
                f'<td style="padding:0.5rem 0.65rem;font-weight:600;color:#0F1B2D;">{meta.get("품명","")}</td>'
                f'<td style="padding:0.5rem 0.65rem;color:#475569;font-size:0.76rem;">{meta.get("카테고리","")}</td>'
                f'<td style="padding:0.5rem 0.65rem;color:#16A34A;font-weight:600;font-family:monospace;">{pos}</td>'
                f'<td style="padding:0.5rem 0.65rem;color:#DC2626;font-weight:700;font-family:monospace;">{neg}</td>'
                f'<td style="padding:0.5rem 0.65rem;color:#B45309;font-weight:600;font-family:monospace;">{neu}</td>'
                f'<td style="padding:0.5rem 0.65rem;">{issues_html}</td>'
                f'</tr>'
            )
        table_html += '</tbody></table></div>'
        st.markdown(table_html, unsafe_allow_html=True)

    st.markdown("---")

    # ────────────────────── (2) 감성분석 정확도 (goldset vs 크롤링 결과)
    st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">AI 판정 정확도 <span style="color:#A0AEC0;font-weight:500;">· 검토 목록 대조</span></div>', unsafe_allow_html=True)

    if not current_results:
        st.caption("먼저 크롤링·분석을 실행하세요.")
    else:
        # goldset URL과 겹치는 결과만 평가
        gs_map = {}
        for r in goldset_rows:
            url = (r.get("url") or "").strip()
            label = (r.get("label") or "").strip()
            if url and label in ("긍정", "부정", "중립"):
                gs_map[url] = label
        # goldset 확정으로 이미 정답 라벨을 쓴 결과는 평가에서 제외 (자기 자신)
        evaluatable = [
            r for r in current_results
            if (r.get("link", "") in gs_map) and (not r.get("goldset"))
        ]
        if not evaluatable:
            st.caption("이번 결과 중 정답셋 URL과 겹치면서 자동 재사용되지 않은 항목이 없어 평가할 수 없습니다.")
        else:
            correct = sum(1 for r in evaluatable if r.get("감성") == gs_map[r["link"]])
            total = len(evaluatable)
            acc = round(correct / total * 100, 1) if total else 0
            # 혼동행렬
            labels = ["긍정", "부정", "중립"]
            cm = {a: {b: 0 for b in labels} for a in labels}
            for r in evaluatable:
                gt = gs_map[r["link"]]
                pd_ = r.get("감성", "중립")
                if gt in labels and pd_ in labels:
                    cm[gt][pd_] += 1
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown(f"""
                <div class="metric-card total">
                    <div class="metric-label">정확도<span class="metric-icon" style="color:#FFFFFF !important;">%</span></div>
                    <div class="metric-value">{acc}%</div>
                    <div class="metric-pct">{correct} / {total} 건 일치</div>
                </div>
                """, unsafe_allow_html=True)
            with c2:
                cm_html = '<div class="card" style="padding:0.4rem 0.6rem;"><table style="width:100%;border-collapse:collapse;font-size:0.82rem;"><thead><tr style="background:#F1F5F9;"><th style="padding:0.4rem 0.6rem;font-size:0.7rem;color:#475569;">정답 \\ 예측</th>'
                for l in labels:
                    cm_html += f'<th style="padding:0.4rem 0.6rem;font-size:0.7rem;color:#475569;text-transform:uppercase;">{l}</th>'
                cm_html += '</tr></thead><tbody>'
                for gt in labels:
                    cm_html += f'<tr style="border-top:1px solid #E1E7F0;"><td style="padding:0.4rem 0.6rem;font-weight:600;color:#0F1B2D;">{gt}</td>'
                    for pd_ in labels:
                        v = cm[gt][pd_]
                        color = "#16A34A" if gt == pd_ and v > 0 else ("#DC2626" if v > 0 else "#A0AEC0")
                        weight = 700 if gt == pd_ else 500
                        cm_html += f'<td style="padding:0.4rem 0.6rem;font-family:monospace;color:{color};font-weight:{weight};text-align:center;">{v}</td>'
                    cm_html += '</tr>'
                cm_html += '</tbody></table></div>'
                st.markdown(cm_html, unsafe_allow_html=True)

    st.markdown("---")

    # ────────────────────── (3) 제외 학습 특징어 채택
    st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">자동 걸러진 광고 단어 <span style="color:#A0AEC0;font-weight:500;">· 제외 라벨링에서 추출</span></div>', unsafe_allow_html=True)

    try:
        gs_map = goldset_lookup_by_url()
    except Exception:
        gs_map = {}
    candidate_kws = _extract_excluded_features(gs_map)

    # 이미 룰베이스에 등록된 것은 제외
    try:
        existing = set(load_keywords_from_sheet().get("exclude", []))
    except Exception:
        existing = set()
    candidate_kws = [k for k in candidate_kws if k not in existing]

    if not candidate_kws:
        st.caption("제외 라벨링이 부족하거나 유의미한 특징 단어가 추출되지 않았습니다. 결과 카드에서 [광고 · 무관으로 제외]를 활용해 주세요.")
    else:
        st.caption(f"후보 {len(candidate_kws)}개. 채택 시 시트 [제외 키워드]에 자동 추가되어 다음 크롤링부터 필터링됩니다.")
        # 뱃지 렌더 + 개별 채택 버튼 (한 줄에 여러 개)
        for i, kw in enumerate(candidate_kws[:20]):
            col_a, col_b, col_c = st.columns([2, 1, 6])
            with col_a:
                st.markdown(f'<div style="padding:0.35rem 0;font-size:0.85rem;font-weight:600;color:#0F1B2D;font-family:\'IBM Plex Sans KR\',sans-serif;">{kw}</div>', unsafe_allow_html=True)
            with col_b:
                if st.button("채택", key=f"adopt_kw_{i}", use_container_width=True, type="secondary"):
                    try:
                        append_keyword_to_sheet("exclude", kw)
                        st.success(f"'{kw}' 제외 키워드에 추가")
                        st.rerun()
                    except Exception as e:
                        st.error(f"저장 실패: {e}")

    # ────────────────────── (4) 정답셋 소급 채우기 (AI 자동)
    st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">기존 정답셋 자동 채우기 <span style="color:#A0AEC0;font-weight:500;">· 상품 · 이슈 비어있는 행에 AI 초안</span></div>', unsafe_allow_html=True)

    # 대상 개수 미리 계산
    _need_count = 0
    for r in goldset_rows:
        if (r.get("label") or "").strip() in ("긍정", "부정", "중립"):
            if not (r.get("product_code") or "").strip() or not (r.get("issue") or "").strip():
                _need_count += 1

    if _need_count == 0:
        st.caption("모든 정답셋 행이 이미 상품/이슈 정보를 갖고 있거나, 채울 대상이 없습니다.")
    else:
        _batches = (_need_count + 19) // 20
        st.markdown(f"""
        <div class="card" style="padding:0.85rem 1rem;margin-bottom:0.75rem;">
            <div style="font-family:'IBM Plex Sans KR',sans-serif;font-size:0.85rem;color:#0F1B2D;line-height:1.6;">
                <b>상품/이슈가 비어있는 정답셋 {_need_count}건</b>을 AI로 자동 채웁니다.
            </div>
            <div style="font-size:0.75rem;color:#4A5568;margin-top:0.35rem;font-family:'IBM Plex Sans KR',sans-serif;line-height:1.7;">
                · 20건씩 묶어 AI 요청 <b>{_batches}회</b> 사용 예정 (한도 100회 중)<br>
                · 채워지면 상품별 고객 이슈 표와 이슈 워드클라우드에 즉시 반영됩니다.<br>
                · 이미 채워진 행은 건너뜁니다.
            </div>
        </div>
        """, unsafe_allow_html=True)

        _bc1, _bc2 = st.columns([1.5, 3.5])
        with _bc1:
            _run = st.button("AI 자동 채우기 시작", key="backfill_run", use_container_width=True)
        with _bc2:
            st.caption("클릭 후 몇 분 대기 (완료까지 페이지를 벗어나지 마세요)")

        if _run:
            ok, reason = _gemini_can_call()
            if not ok:
                st.warning(f"AI 호출 불가 · {reason}")
            else:
                progress = st.progress(0)
                status = st.empty()

                def _cb(done, total):
                    try:
                        progress.progress(min(done / max(total, 1), 1.0))
                        status.markdown(
                            f'<span style="font-size:0.78rem;color:#7A879E;font-family:\'IBM Plex Mono\',monospace;">진행 {done} / {total}</span>',
                            unsafe_allow_html=True
                        )
                    except Exception:
                        pass

                with st.spinner("AI가 상품·이슈 채우는 중..."):
                    res = backfill_goldset_products_issues(progress_callback=_cb)
                progress.empty()
                status.empty()

                _msg = (
                    f"처리 {res['processed']} · 업데이트 {res['updated']} · 건너뜀 {res['skipped']}"
                )
                if res.get("stopped_reason"):
                    st.warning(f"{_msg}\n\n사유 · {res['stopped_reason']}")
                else:
                    st.success(_msg)


# ============================================== AI 라벨링 (관리자 인라인 UI)
def _run_batch_ai_labeling(page_results: list, src_name: str, page: int):
    """페이지의 20개 결과를 한 번에 Gemini에 넘겨 초안 라벨을 세션에 저장."""
    ok, reason = _gemini_can_call()
    if not ok:
        st.warning(f"Gemini 호출 불가: {reason}")
        return
    items_for_ai = []
    for idx, r in enumerate(page_results):
        # 이미 정답셋 확정된 항목은 스킵
        if r.get("goldset"):
            continue
        items_for_ai.append({
            "idx": idx,
            "title": r.get("title", ""),
            "description": r.get("description", "") or (r.get("reason", "") or ""),
        })
    if not items_for_ai:
        st.info("일괄 처리할 항목이 없습니다.")
        return
    # 텍스트 합쳐서 후보 상품 추림
    combined = " ".join((it["title"] + " " + it["description"]) for it in items_for_ai)[:3000]
    candidates = match_all_products(combined, max_hits=15)
    with st.spinner(f"AI 분석 중... ({len(items_for_ai)}건 · 한 요청)"):
        results = analyze_batch_with_gemini(items_for_ai, candidates)
    if not results:
        st.warning("배치 분석 실패. 개별 카드의 [AI로 상품 · 이슈 뽑아보기]로 시도해 보세요.")
        return
    # 각 결과를 세션에 반영
    saved = 0
    for res in results:
        try:
            idx = int(res.get("idx"))
        except Exception:
            continue
        if idx < 0 or idx >= len(page_results):
            continue
        item = page_results[idx]
        key_base = f"ai_{src_name}_{page}_{idx}"
        state_key = f"{key_base}_state"
        st.session_state[state_key] = {
            "label": res.get("label") or item.get("감성", "중립"),
            "product_code": res.get("품번") or (item.get("품번", "").split(",")[0].strip() if item.get("품번") else ""),
            "product_name": res.get("품명") or item.get("품명", ""),
            "issue": res.get("issue") or "",
            "generated": True,
        }
        saved += 1
    st.success(f"{saved}건 AI 초안 생성. 각 카드 [확인 · 저장]에서 검수하세요.")


def _render_ai_labeler(item: dict, src_name: str, page: int, idx: int):
    """관리자 모드에서 각 결과 카드 아래에 접이식 AI 라벨링 UI를 표시.
    - [AI로 상품·이슈 뽑아보기] 버튼: Gemini 호출해 초안 생성
    - 편집 필드: 라벨/상품(품번)/이슈 수정 가능
    - [저장하기] 버튼: goldset 시트에 append
    - [광고·무관으로 제외] 버튼: label=제외로 저장
    """
    url = item.get("link", "")
    key_base = f"ai_{src_name}_{page}_{idx}"
    state_key = f"{key_base}_state"   # {label, product_code, product_name, issue, generated}

    # 기본 초기값 (룰베이스 결과에서 가져옴)
    if state_key not in st.session_state:
        st.session_state[state_key] = {
            "label": item.get("감성", "중립"),
            "product_code": item.get("품번", "").split(",")[0].strip() if item.get("품번") else "",
            "product_name": item.get("품명", "") or "",
            "issue": "",
            "generated": False,
        }
    state = st.session_state[state_key]

    with st.expander("확인 · 저장", expanded=False):
        c1, c2, c3 = st.columns([1.3, 1, 1])
        with c1:
            if st.button("AI로 상품 · 이슈 뽑아보기", key=f"{key_base}_run",
                         use_container_width=True, type="secondary"):
                ok, reason = _gemini_can_call()
                if not ok:
                    st.warning(f"Gemini 호출 불가: {reason}")
                else:
                    with st.spinner("Gemini 분석 중..."):
                        text_for_ai = (item.get("title", "") + " \n " + item.get("description", ""))[:2800]
                        # 품명DB에서 텍스트에 등장하는 후보만 우선 추림
                        candidates = match_all_products(text_for_ai, max_hits=15)
                        if not candidates and not PRODUCT_DB.empty:
                            candidates = [
                                {"품번": str(row.get("품번", "")).strip(),
                                 "품명": str(row.get("품명", "")).strip(),
                                 "소분류": str(row.get("소분류", "")).strip()}
                                for _, row in PRODUCT_DB.head(30).iterrows()
                            ]
                        results = analyze_with_gemini(text_for_ai, candidates)
                        if results:
                            first = results[0]
                            state["label"] = first.get("label") or state["label"]
                            state["product_code"] = first.get("품번") or state["product_code"]
                            state["product_name"] = first.get("품명") or state["product_name"]
                            state["issue"] = first.get("issue") or ""
                            state["generated"] = True
                            # 결과가 여러 개면 세션에 저장 (참고 표시용)
                            if len(results) > 1:
                                state["extra_hits"] = results[1:]
                            st.rerun()
                        else:
                            st.info("AI 분석 결과가 비어있습니다. 다이소 상품 언급이 없거나 인식 불가.")
        with c2:
            new_label = st.selectbox(
                "라벨", ["긍정", "부정", "중립", "제외"],
                index=["긍정", "부정", "중립", "제외"].index(state["label"]) if state["label"] in ["긍정","부정","중립","제외"] else 2,
                key=f"{key_base}_label", label_visibility="collapsed",
            )
            state["label"] = new_label
        with c3:
            state["product_code"] = st.text_input(
                "품번", value=state["product_code"], key=f"{key_base}_code",
                label_visibility="collapsed", placeholder="품번",
            )

        # 상품명 + 이슈 (한 줄)
        c4, c5 = st.columns([1, 1.4])
        with c4:
            state["product_name"] = st.text_input(
                "상품명", value=state["product_name"], key=f"{key_base}_pname",
                label_visibility="collapsed", placeholder="상품명 (자동 조인)",
            )
        with c5:
            state["issue"] = st.text_input(
                "이슈 요약", value=state["issue"], key=f"{key_base}_issue",
                label_visibility="collapsed", placeholder="이슈 한 줄 (예: 뚜껑 뻑뻑함)",
            )

        # AI가 추가로 뽑은 상품이 있으면 참고 표시
        extras = state.get("extra_hits") or []
        if extras:
            hint = " · ".join(f"{h.get('품명','')}({h.get('label','')}: {h.get('issue','')})" for h in extras[:3])
            st.caption(f"AI 추가 감지 · {hint}")

        # 저장/제외 버튼
        c6, c7, _sp = st.columns([1, 1, 3])
        with c6:
            if st.button("저장하기", key=f"{key_base}_save", use_container_width=True):
                text_for_save = (item.get("title", "") + " " + item.get("description", ""))[:500]
                append_goldset_to_sheet(
                    url=url,
                    title=item.get("title", ""),
                    label=state["label"] or "중립",
                    text_snippet=text_for_save,
                    product_code=state["product_code"] or "",
                    issue=state["issue"] or "",
                )
                st.success("정답셋에 저장했습니다.")
        with c7:
            if st.button("광고 · 무관으로 제외", key=f"{key_base}_exclude", use_container_width=True, type="secondary"):
                append_goldset_to_sheet(
                    url=url,
                    title=item.get("title", ""),
                    label="제외",
                    text_snippet=(item.get("title", "") + " " + item.get("description", ""))[:500],
                    product_code="",
                    issue=state["issue"] or "광고/무관",
                )
                append_excluded_url_to_sheet(url, reason=state["issue"] or "관리자 제외")
                st.success("제외 처리 완료 (다음 크롤링부터 필터링).")


def render_detail_tab(src_results, src_name, start_date, end_date):
    if not src_results:
        st.info(f"{src_name} 수집 결과가 없습니다."); return
    t  = len(src_results)
    p  = sum(1 for r in src_results if r["감성"]=="긍정")
    n  = sum(1 for r in src_results if r["감성"]=="부정")
    ne = sum(1 for r in src_results if r["감성"]=="중립")

    # ── 메트릭 카드 (표시 전용) ──
    c1, c2, c3, c4 = st.columns(4)
    for col, cls, lbl, val, ic_txt in [
        (c1,"total","전체",str(t),"전체"),
        (c2,"pos","긍정",str(p),"긍정"),
        (c3,"neg","부정",str(n),"부정"),
        (c4,"neu","중립",str(ne),"중립"),
    ]:
        with col:
            st.markdown(f"""
            <div class="metric-card {cls}">
                <div class="metric-label">
                    <span class="metric-icon {cls}" style="color:#FFFFFF !important;">{ic_txt}</span>{lbl}
                </div>
                <div class="metric-value">{val}</div>
                <div class="metric-pct">{round(int(val)/t*100) if t else 0}%</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:0.75rem'></div>", unsafe_allow_html=True)

    # ── 필터(왼쪽) + 정렬(오른쪽) ──
    filter_col, _, sort_col = st.columns([3, 4, 4])
    with filter_col:
        current_filter = st.radio("감성 필터", ["전체", "긍정", "부정", "중립"],key=f"filter_{src_name}", horizontal=True, label_visibility="collapsed")
    with sort_col:
        sort_opt = st.radio("정렬", ["관련성 낮은순", "관련성 높은순", "부정 높은순", "부정 낮은순", "최신순", "오래된순"],key=f"sort_{src_name}", horizontal=True, label_visibility="collapsed")

    # ── 필터 적용 ──
    if current_filter == "긍정":
        src_results = [r for r in src_results if r["감성"] == "긍정"]
    elif current_filter == "부정":
        src_results = [r for r in src_results if r["감성"] == "부정"]
    elif current_filter == "중립":
        src_results = [r for r in src_results if r["감성"] == "중립"]

    # ── 정렬 적용 ──
    if sort_opt == "관련성 낮은순":
        # 노이즈가 앞으로 몰림 → 관리자 대량 제외 학습에 유리
        src_results = sorted(src_results, key=lambda x: x.get("관련성", 0))
    elif sort_opt == "관련성 높은순":
        src_results = sorted(src_results, key=lambda x: x.get("관련성", 0), reverse=True)
    elif sort_opt == "부정 높은순":
        src_results = sorted(src_results, key=lambda x: x.get("확신도", 0) if x.get("감성") == "부정" else 0, reverse=True)
    elif sort_opt == "부정 낮은순":
        src_results = sorted(src_results, key=lambda x: x.get("확신도", 0) if x.get("감성") == "부정" else 100)
    elif sort_opt == "최신순":
        src_results = sorted(src_results, key=lambda x: x.get("날짜", ""), reverse=True)
    elif sort_opt == "오래된순":
        src_results = sorted(src_results, key=lambda x: x.get("날짜", ""))

    # ── 상세 결과 헤더 ──
    st.markdown(f'<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">상세 결과 <span style="color:#A0AEC0;font-weight:500;">· {len(src_results)}건</span></div>', unsafe_allow_html=True)

    PAGE_SIZE = 20
    total_pages = (len(src_results) - 1) // PAGE_SIZE + 1 if src_results else 1
    page_key = f"page_{src_name}"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    current_page = st.session_state[page_key]

    start_idx = (current_page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_results = src_results[start_idx:end_idx]

    # ── 골드셋 버튼 (관리자, 오른쪽 상단 소형) ──
    if st.session_state.get("admin_mode"):
        checked_items = [(i, page_results[i]) for i in range(len(page_results))
                        if st.session_state.get(f"chk_{src_name}_{current_page}_{i}")]
        checked_urls = [item["link"] for _, item in checked_items]
        _sp, _bt0, _g1, _g2, _g3 = st.columns([3.5, 2, 1.5, 1.5, 1.5])
        with _bt0:
            if st.button("이 페이지 20개 한 번에 뽑기", key=f"bulk_ai_{src_name}_{current_page}", use_container_width=True, type="secondary", disabled=not gemini_available()):
                _run_batch_ai_labeling(page_results, src_name, current_page)
                st.rerun()
        with _g1:
            if st.button(f"긍정 저장 ({len(checked_urls)})", key=f"gold_pos_{src_name}_{current_page}", disabled=len(checked_urls)==0, use_container_width=True, type="secondary"):
                for _, item in checked_items:
                    append_goldset_to_sheet(item["link"], item.get("title",""), "긍정", clean_text(item.get("title","")+" "+item.get("link","")))
                st.success(f"{len(checked_urls)}건 긍정 저장")
                st.rerun()
        with _g2:
            if st.button(f"부정 저장 ({len(checked_urls)})", key=f"gold_neg_{src_name}_{current_page}", disabled=len(checked_urls)==0, use_container_width=True, type="secondary"):
                for _, item in checked_items:
                    append_goldset_to_sheet(item["link"], item.get("title",""), "부정", clean_text(item.get("title","")+" "+item.get("link","")))
                st.success(f"{len(checked_urls)}건 부정 저장")
                st.rerun()
        with _g3:
            if st.button(f"제외 저장 ({len(checked_urls)})", key=f"bulk_exc_{src_name}_{current_page}", disabled=len(checked_urls)==0, use_container_width=True, type="secondary"):
                for url in checked_urls:
                    append_excluded_url_to_sheet(url, reason="관리자 일괄 제외")
                st.session_state["analysis_results"] = [
                    r for r in st.session_state["analysis_results"] if r.get("link") not in checked_urls
                ]
                st.success(f"{len(checked_urls)}건 제외 완료")
                st.rerun()

    for idx, r in enumerate(page_results):
        _b     = SENT_BADGE.get(r["감성"], "")
        _sub   = ('<span class="badge-sub">카테고리 · ' + r["소분류"] + '</span>') if r.get("소분류")   else ""
        _code  = ('<span class="badge-sub">품번 · ' + r["품번"] + '</span>') if r.get("품번")     else ""
        _name  = ('<span>품명 · ' + r["품명"] + '</span>') if r.get("품명")     else ""
        _price = ('<span>가격 · ' + r["가격언급"] + '</span>') if r.get("가격언급") else ""
        _gs    = ('<span style="background:#0F1B2D;color:#FFFFFF;padding:2px 8px;border-radius:2px;font-size:0.68rem;font-weight:700;letter-spacing:0.06em;text-transform:uppercase;">검토 완료</span>') if r.get("goldset") else ""
        _issue = ('<span style="color:#1E3A8A;font-weight:600;">요점 · ' + r["issue"] + '</span>') if r.get("issue") else ""
        _badge = '<span class="' + _b + '">' + r["감성"] + ' ' + fmt_score(r["확신도"]) + '</span>'
        _rel_score = r.get("관련성", 0)
        if _rel_score >= 99:
            _rel = ''  # 정답셋 확정은 별도 뱃지 있음
        elif _rel_score >= 15:
            _rel = f'<span style="background:#EFF4FF;color:#1D4ED8;padding:2px 8px;border-radius:2px;font-size:0.68rem;font-weight:700;letter-spacing:0.04em;text-transform:uppercase;">관련도 상 · {_rel_score}</span>'
        elif _rel_score >= 8:
            _rel = f'<span style="background:#F1F5F9;color:#475569;padding:2px 8px;border-radius:2px;font-size:0.68rem;font-weight:600;letter-spacing:0.04em;text-transform:uppercase;">관련도 중 · {_rel_score}</span>'
        else:
            _rel = f'<span style="background:#FEF2F2;color:#DC2626;padding:2px 8px;border-radius:2px;font-size:0.68rem;font-weight:600;letter-spacing:0.04em;text-transform:uppercase;">관련도 저 · {_rel_score}</span>'
        _reason = ('<div style="font-size:0.75rem;color:#4A5568;margin-top:0.4rem;padding-left:0.7rem;border-left:2px solid #CBD5E1;font-family:\'IBM Plex Sans KR\',sans-serif;"><span style="font-size:0.62rem;color:#7A879E;font-family:\'IBM Plex Mono\',monospace;text-transform:uppercase;letter-spacing:0.05em;">근거 </span>' + r["reason"] + '</div>') if r.get("reason") else ""
        _title = r["title"] or "(제목 없음)"
        if st.session_state.get("admin_mode"):
            col_chk, col_card = st.columns([0.3, 9.7])
            with col_chk:
                st.checkbox("", key=f"chk_{src_name}_{current_page}_{idx}", label_visibility="collapsed")
            with col_card:
                st.markdown(
                    '<div class="result-card"><div class="result-title">'
                    '<a href="' + r["link"] + '" target="_blank" style="color:#1A202C;text-decoration:none;">' + _title + '</a>'
                    + ' ' + _gs + ' ' + _rel +
                    '</div><div class="result-meta">'
                    '<span>검색어 · ' + r["검색어"] + '</span><span>날짜 · ' + r["날짜"] + '</span>'
                    + _sub + _code + _name + _price + _issue + _badge +
                    '</div>' + _reason + '</div>', unsafe_allow_html=True)
                _render_ai_labeler(r, src_name, current_page, idx)
        else:
            st.markdown(
                '<div class="result-card"><div class="result-title">'
                '<a href="' + r["link"] + '" target="_blank" style="color:#1A202C;text-decoration:none;">' + _title + '</a>'
                + ' ' + _gs + ' ' + _rel +
                '</div><div class="result-meta">'
                '<span>검색어 · ' + r["검색어"] + '</span><span>날짜 · ' + r["날짜"] + '</span>'
                + _sub + _code + _name + _price + _issue + _badge +
                '</div>' + _reason + '</div>', unsafe_allow_html=True)

    if total_pages > 1:
        pg_col1, pg_col2, pg_col3 = st.columns([1, 2, 1])
        with pg_col1:
            if st.button("이전", key=f"prev_{src_name}", disabled=(current_page <= 1)):
                st.session_state[page_key] = current_page - 1
                st.rerun()
        with pg_col2:
            st.markdown(f'<div style="text-align:center;font-size:0.85rem;color:#4A5568;padding:0.5rem;">{current_page} / {total_pages} 페이지</div>', unsafe_allow_html=True)
        with pg_col3:
            if st.button("다음", key=f"next_{src_name}", disabled=(current_page >= total_pages)):
                st.session_state[page_key] = current_page + 1
                st.rerun()

    src_csv = pd.DataFrame(src_results).to_csv(index=False, encoding="utf-8-sig")
    st.download_button(f"{src_name} 전체 CSV 다운로드 ({len(src_results)}건)", src_csv.encode("utf-8-sig"),
        f"ISSUE_{src_name}_{start_date}_{end_date}.csv", "text/csv", use_container_width=True)

# ============================
# 관리자 모드 토글 (role=admin만 표시)
# ============================
admin_col1, admin_col2 = st.columns([10, 1])
with admin_col2:
    if st.session_state.get("current_user_role") == "admin":
        if st.session_state["admin_mode"]:
            if st.button("관리자 OFF", key="admin_toggle_off"):
                st.session_state["admin_mode"] = False
                st.rerun()
            st.markdown('<span class="admin-badge-on">ADMIN</span>', unsafe_allow_html=True)
        else:
            if st.button("관리자 ON", key="admin_toggle_on"):
                st.session_state["admin_mode"] = True
                st.rerun()

# ============================
# 앱 헤더
# ============================
st.markdown("""
<div class="app-header">
    <div style="display:flex;align-items:center;gap:0.5rem;flex-shrink:0;">
        <div style="
            width:48px; height:48px;
            background:#0066CC;
            border-radius:50%;
            display:flex; align-items:center; justify-content:center;
            flex-shrink:0;
            box-shadow:0 2px 6px rgba(0,102,204,0.35);
        ">
            <svg width="30" height="20" viewBox="0 0 60 38" fill="none" xmlns="http://www.w3.org/2000/svg">
                <path d="M0 2 H8 Q16 2 16 10 Q16 18 8 18 H0 Z M4 5 V15 H8 Q12 15 12 10 Q12 5 8 5 Z" fill="#FFFFFF"/>
                <path d="M18 18 L24 2 L30 18 M20.5 12 H27.5" stroke="#FFFFFF" stroke-width="3" fill="none" stroke-linecap="round"/>
                <rect x="33" y="2" width="3.5" height="16" rx="1" fill="#FFFFFF"/>
                <path d="M40 15 Q40 18 44 18 Q48 18 48 14.5 Q48 11 44 10 Q40 9 40 5.5 Q40 2 44 2 Q48 2 48 5"
                      stroke="#FFFFFF" stroke-width="3" fill="none" stroke-linecap="round"/>
                <ellipse cx="54" cy="10" rx="5" ry="8" stroke="#FFFFFF" stroke-width="3" fill="none"/>
            </svg>
        </div>
        <div style="
            font-size:1.35rem; font-weight:900;
            color:#0066CC; letter-spacing:0.12em;
            font-family:'Inter',sans-serif;
            line-height:1;
        "></div>
    </div>
    <div style="width:1px;height:36px;background:#E2E8F0;margin:0 0.25rem;flex-shrink:0;"></div>
    <div>
        <div class="header-title"> SNS 고객 불만 AI분석 Platform</div>
        <div class="header-sub">네이버 블로그 · 카페 · 유튜브 &nbsp;|&nbsp; KLUE-RoBERTa + 룰베이스 앙상블</div>
    </div>
</div>
""", unsafe_allow_html=True)


# ============================
# 사이드바
# ============================
with st.sidebar:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:0.6rem;padding-bottom:1rem;border-bottom:1px solid #E2E8F0;margin-bottom:0.25rem;">
        <div style="width:32px;height:32px;background:#0066CC;border-radius:50%;display:flex;align-items:center;justify-content:center;box-shadow:0 1px 4px rgba(0,102,204,0.3);">
            <span style="color:#FFFFFF;font-size:0.65rem;font-weight:900;letter-spacing:0.05em;font-family:'Inter',sans-serif;">D</span>
        </div>
        <div>
            <div style="font-weight:700;font-size:0.95rem;color:#1A202C;">DAISO ISSUE FINDER</div>
            <div style="font-size:0.68rem;color:#718096;">Created by 데이터분석팀</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.4rem;">
        <div class="sb-section-icon">
            <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <path d="M4 11a9 9 0 0 1 9 9"/><path d="M4 4a16 16 0 0 1 16 16"/><circle cx="5" cy="19" r="1"/>
            </svg>
        </div>
        <span class="sb-section-text">SOURCE</span>
    </div>
    """, unsafe_allow_html=True)

    row1_left, row1_right = st.columns(2)
    with row1_left:
        cb_col, icon_col = st.columns([1, 4])
        with cb_col:
            search_blog = st.checkbox("", value=True, key="cb_blog", label_visibility="collapsed")
        with icon_col:
            st.markdown("""<div class="ch-row">
                <div class="ch-icon ch-naver">N</div>
                <span class="ch-label">블로그</span>
            </div>""", unsafe_allow_html=True)

    with row1_right:
        cb_col2, icon_col2 = st.columns([1, 4])
        with cb_col2:
            search_cafe = st.checkbox("", value=True, key="cb_cafe", label_visibility="collapsed")
        with icon_col2:
            st.markdown("""<div class="ch-row">
                <div class="ch-icon ch-naver">N</div>
                <span class="ch-label">카페</span>
            </div>""", unsafe_allow_html=True)

    row2_left, row2_right = st.columns(2)
    with row2_left:
        cb_col_kin, icon_col_kin = st.columns([1, 4])
        with cb_col_kin:
            search_kin = st.checkbox("", value=True, key="cb_kin", label_visibility="collapsed")
        with icon_col_kin:
            st.markdown("""<div class="ch-row">
                <div class="ch-icon ch-naver">N</div>
                <span class="ch-label">지식iN</span>
            </div>""", unsafe_allow_html=True)

    with row2_right:
        cb_col3, icon_col3 = st.columns([1, 4])
        with cb_col3:
            search_yt = st.checkbox("", value=True, key="cb_yt", label_visibility="collapsed")
        with icon_col3:
            st.markdown("""<div class="ch-row">
                <div class="ch-icon ch-youtube">
                    <svg width="9" height="9" viewBox="0 0 24 24" fill="#FFFFFF"><polygon points="5,3 19,12 5,21"/></svg>
                </div>
                <span class="ch-label">유튜브</span>
            </div>""", unsafe_allow_html=True)

    row3_left, row3_right = st.columns(2)
    with row3_left:
        cb_col_yc, icon_col_yc = st.columns([1, 4])
        with cb_col_yc:
            search_yt_comments = st.checkbox("", value=True, key="cb_yt_comments", label_visibility="collapsed")
        with icon_col_yc:
            st.markdown("""<div class="ch-row">
                <div class="ch-icon ch-youtube">
                    <svg width="9" height="9" viewBox="0 0 24 24" fill="#FFFFFF"><polygon points="5,3 19,12 5,21"/></svg>
                </div>
                <span class="ch-label">유튜브 댓글</span>
            </div>""", unsafe_allow_html=True)

    with row3_right:
        cb_col_pp, icon_col_pp = st.columns([1, 4])
        with cb_col_pp:
            search_ppomppu = st.checkbox("", value=False, key="cb_ppomppu", label_visibility="collapsed")
        with icon_col_pp:
            st.markdown("""<div class="ch-row">
                <div style="width:20px;height:20px;border-radius:4px;background:#6366F1;display:flex;align-items:center;justify-content:center;font-size:0.6rem;color:#FFFFFF;font-weight:900;flex-shrink:0;">C</div>
                <span class="ch-label">커뮤니티</span>
            </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.3rem;">
        <div class="sb-section-icon">
            <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/>
            </svg>
        </div>
        <span class="sb-section-text">분석 검색어</span>
    </div>
    """, unsafe_allow_html=True)
    keywords_input = st.text_area("", value="다이소 상품불량\n다이소 불량\n다이소 별로",
                                  height=95, label_visibility="collapsed",
                                  placeholder="줄바꿈으로 구분 · 최대 3개")
    st.markdown('<span class="sb-hint">줄바꿈으로 구분, 최대 3개<br>※ \'다이소\' 없으면 자동 추가됩니다</span>', unsafe_allow_html=True)

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.3rem;">
        <div class="sb-section-icon">
            <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <rect x="3" y="4" width="18" height="18" rx="2" ry="2"/>
                <line x1="16" y1="2" x2="16" y2="6"/><line x1="8" y1="2" x2="8" y2="6"/>
                <line x1="3" y1="10" x2="21" y2="10"/>
            </svg>
        </div>
        <span class="sb-section-text">분석 기간</span>
    </div>
    """, unsafe_allow_html=True)

    dc1, dc2 = st.columns(2, gap="small")
    with dc1:
        st.markdown('<span class="date-label">시작일</span>', unsafe_allow_html=True)
        start_date = st.date_input("시작일", value=date(2026, 1, 1), label_visibility="collapsed", key="date_start")
    with dc2:
        st.markdown('<span class="date-label">종료일</span>', unsafe_allow_html=True)
        end_date = st.date_input("종료일", value=date.today(), label_visibility="collapsed", key="date_end")

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.3rem;">
        <div class="sb-section-icon">
            <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/>
                <line x1="8" y1="18" x2="21" y2="18"/><line x1="3" y1="6" x2="3.01" y2="6"/>
                <line x1="3" y1="12" x2="3.01" y2="12"/><line x1="3" y1="18" x2="3.01" y2="18"/>
            </svg>
        </div>
        <span class="sb-section-text">분석개수</span>
    </div>
    """, unsafe_allow_html=True)
    display_count = st.number_input(
        "", min_value=50, max_value=1000, value=100, step=50,
        label_visibility="collapsed",
        help="CPU 환경 권장 수집건수 (최소 50 ~ 최대 1,000)"
    )
    st.markdown('<span class="sb-hint">CPU 권장: 100~300건 · 최대 1,000건</span>', unsafe_allow_html=True)

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.3rem;">
        <div class="sb-section-icon">
            <svg width="11" height="11" viewBox="0 0 24 24" fill="none" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                <path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/>
                <polyline points="22 4 12 14.01 9 11.01"/>
            </svg>
        </div>
        <span class="sb-section-text">감성 파라미터</span>
    </div>
    """, unsafe_allow_html=True)
    threshold = st.number_input(
        "", min_value=40, max_value=95, value=55, step=5,
        label_visibility="collapsed",
        help="AI가 이 수치 이상의 확신도로 부정 판정 시에만 부정으로 등록"
    )

    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        run_btn = st.button("분석 시작", use_container_width=True)
    with btn_col2:
        stop_btn = st.button("중지", use_container_width=True)

    st.markdown("""
    <div class="sb-section" style="margin:0.5rem 0 0.3rem;">
        <div class="sb-section-icon">S</div>
        <span class="sb-section-text">감성 파라미터 가이드</span>
    </div>
    <div class="param-guide-box">
        <b>고객 반응 판정 강도 조정</b><br>
        • <code>40~50%</code> → 민감하게 수집 (부정 많이 잡힘)<br>
        • <code>55~65%</code> → 권장 (정확도 균형)<br>
        • <code>70%+</code> → 엄격 (확실한 부정만)<br><br>
        <b>단어 직접 추가</b><br>
        관리자 모드에서 [긍정/부정/홍보성] 단어를 추가하면 해당 단어가 포함된 글을 처리합니다.<br><br>
        <b>현재 AI 모델 가중치</b><br>
        • KLUE-RoBERTa 가중치: <code>* 1.8</code> (메인 모델)<br>
        • Rule-Base: <code>* 1.5</code> (키워드 보강)<br>
    </div>
    """, unsafe_allow_html=True)


# ============================
# 분석 실행
# ============================
if stop_btn:
    st.session_state.pop("analysis_results", None)
    st.warning("분석이 중지되었습니다.")
    st.stop()

if run_btn:
    keywords_raw = [k.strip() for k in keywords_input.strip().splitlines() if k.strip()][:3]
    if not keywords_raw:
        st.error("검색어를 최소 1개 입력해주세요."); st.stop()
    if not any([search_blog, search_cafe, search_kin, search_yt, search_yt_comments, search_ppomppu]):
        st.error("채널을 하나 이상 선택해주세요."); st.stop()
    if start_date > end_date:
        st.error("시작일이 종료일보다 늦습니다. 날짜를 확인해주세요."); st.stop()

    # 다중 사용자 환경 대비: 크롤링 시작 시점에 시트에서 최신 필터 재로드
    try:
        load_excluded_urls_from_sheet.clear()
    except Exception:
        pass
    EXCLUDED_URLS_FROM_SHEET = load_excluded_urls_from_sheet()
    try:
        load_keywords_from_sheet.clear()
    except Exception:
        pass
    # 학습된 제외 특징어 캐시도 무효화 (다른 관리자가 방금 제외 라벨링했을 수 있음)
    st.session_state.pop("_learned_exclude_cache", None)

    keywords = [build_naver_query(k) for k in keywords_raw]

    with st.spinner("감성 분석 준비 중..."):
        model_r = load_roberta()

    if model_r is None:
        st.markdown(
            '<div style="background:#EFF4FF;border:1px solid #C7D6F0;border-radius:6px;'
            'padding:0.6rem 1rem;font-size:0.82rem;color:#1E3A8A;font-weight:500;margin-bottom:0.5rem;font-family:\'IBM Plex Sans KR\',sans-serif;">'
            '감성 초벌 판정은 단어 규칙 기반, 정밀 판정은 관리자가 [AI로 상품 · 이슈 뽑아보기]로 처리합니다.</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div style="background:#F0FDF4;border:1px solid #A7F3D0;border-radius:6px;'
            'padding:0.6rem 1rem;font-size:0.82rem;color:#16A34A;font-weight:500;margin-bottom:0.5rem;font-family:\'IBM Plex Sans KR\',sans-serif;">'
            'AI 모델 준비 완료</div>',
            unsafe_allow_html=True
        )

    collect_tasks = []
    for kw in keywords:
        if search_blog: collect_tasks.append(("blog", kw, "블로그"))
        if search_cafe: collect_tasks.append(("cafe", kw, "카페"))
        if search_kin:  collect_tasks.append(("kin", kw, "지식iN"))
        if search_yt and YOUTUBE_API_KEY:
            collect_tasks.append(("yt", kw, "유튜브"))
        if search_yt_comments and YOUTUBE_API_KEY:
            collect_tasks.append(("yt_comments", kw, "유튜브 댓글"))
        if search_ppomppu:
            collect_tasks.append(("ppomppu", kw, "커뮤니티"))

    prog      = st.progress(0)
    prog_text = st.empty()
    all_items = []; collect_log = []


    total_tasks = len(collect_tasks)
    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(_fetch, t, display_count): t for t in collect_tasks}
        for fut in concurrent.futures.as_completed(futures):
            label, kw, items = fut.result()
            all_items.extend(items)
            collect_log.append(f"{label}/{kw}/{len(items)}건")
            done += 1
            prog.progress(done / max(total_tasks, 1))
            prog_text.markdown(f'<span style="font-size:0.78rem;color:#718096;">수집 중 {done}/{total_tasks} 완료</span>', unsafe_allow_html=True)

    prog.empty(); prog_text.empty()

    seen, unique_items = set(), []
    for item in all_items:
        lnk = item.get("link","")
        if lnk not in seen: seen.add(lnk); unique_items.append(item)

    before_rel = len(unique_items)
    unique_items = [it for it in unique_items if is_daiso_related(it)]
    rel_excluded = before_rel - len(unique_items)

    before_promo = len(unique_items)
    unique_items = [it for it in unique_items if not is_promotional(it)]
    promo_excluded = before_promo - len(unique_items)

    unique_items = [it for it in unique_items if not is_admin_excluded(it)]


    before_usim  = len(unique_items)
    unique_items = [it for it in unique_items if not is_usim_related(it)]
    usim_excluded = before_usim - len(unique_items)

    filtered = filter_by_date(unique_items, start_date, end_date)
    if not filtered:
        st.warning("해당 기간에 결과가 없습니다. 날짜 범위나 검색어를 확인해주세요."); st.stop()

    notes = []
    if rel_excluded > 0:    notes.append(f"다이소 무관 <strong>{rel_excluded}</strong>건 제외")
    if promo_excluded > 0:  notes.append(f"홍보성 글 <strong>{promo_excluded}</strong>건 제외")
    if usim_excluded > 0:   notes.append(f"유심 관련 <strong>{usim_excluded}</strong>건 제외")
    note_str = " &nbsp;·&nbsp; ".join(notes)
    if note_str: note_str = " &nbsp;·&nbsp; " + note_str

    st.markdown(f"""
    <div class="card" style="border-left:3px solid #0066CC;">
        <span style="font-size:0.85rem;color:#0066CC;font-weight:600;">
        수집 완료 · 총 <strong>{len(filtered)}</strong>건 (중복 제거 후){note_str}
        </span><br>
        <span style="font-size:0.72rem;color:#718096;">{' &nbsp;|&nbsp; '.join(collect_log)}</span>
    </div>
    """, unsafe_allow_html=True)

    results = []
    progress_bar = st.progress(0)
    status_text  = st.empty()

    BATCH   = 32
    total_f = len(filtered)

    # ── 정답셋 URL→라벨 맵 로드 (재사용 가속) ──
    _goldset_map = {}
    _goldset_excluded_urls = set()
    try:
        _goldset_map = goldset_lookup_by_url()
        _goldset_excluded_urls = {u for u, v in _goldset_map.items() if v.get("label") == "제외"}
    except Exception:
        pass

    # ── 정답셋 [제외] 라벨에서 자동 학습된 특징 키워드 (URL이 아닌 새 글도 필터) ──
    try:
        _learned_exclude_kws = _extract_excluded_features(_goldset_map)
    except Exception:
        _learned_exclude_kws = []

    for batch_start in range(0, total_f, BATCH):
        batch = filtered[batch_start: batch_start + BATCH]
        texts, metas = [], []

        for i, item in enumerate(batch):
            src   = item.get("출처","")
            title = clean_text(item.get("title",""))
            desc  = clean_text(item.get("description",""))
            full = (title + " " + desc)[:300]
            texts.append(full)
            metas.append((src, item, title))

        r_batch = model_r(texts, batch_size=BATCH, truncation=True, max_length=192) if model_r else [None]*len(texts)

        for idx, (full, (src, item, title)) in enumerate(zip(texts, metas)):
            url = item.get("link", "")
            # ── ① 정답셋에 이미 라벨된 URL이면 정답 그대로 사용 (재분석 스킵) ──
            gs_hit = _goldset_map.get(url)
            if gs_hit:
                gs_label = gs_hit.get("label", "")
                if gs_label == "제외":
                    # 관리자가 제외 표시한 글은 결과에서 아예 제거
                    continue
                if gs_label in ("긍정", "부정", "중립"):
                    date_str = item.get("날짜","") if src == "유튜브" else (
                        lambda dt: dt.strftime("%Y-%m-%d") if dt else ""
                    )(parse_date(item))
                    results.append({
                        "출처":    src,
                        "검색어":  item.get("검색어",""),
                        "소분류":  gs_hit.get("category") or "",
                        "품번":    gs_hit.get("product_code") or "",
                        "품명":    gs_hit.get("product_name") or "",
                        "가격언급": "",
                        "title":  title,
                        "link":   url,
                        "날짜":   date_str,
                        "감성":   gs_label,
                        "확신도": 100,
                        "channel": item.get("channel","") or item.get("cafename",""),
                        "views":   item.get("views",""),
                        "likes":   item.get("likes",""),
                        "comments":item.get("comments",""),
                        "video_id":item.get("video_id",""),
                        "reason":  gs_hit.get("issue") or "정답셋 확정",
                        "goldset": True,
                        "issue":   gs_hit.get("issue") or "",
                        "관련성":  99,   # 정답셋 확정은 최고 우선순위
                        "description": item.get("description", ""),
                    })
                    continue

            # ── ② [제외] 학습 키워드가 본문에 다수 등장하면 자동 제외 ──
            if _learned_exclude_kws:
                hit_cnt = sum(1 for kw in _learned_exclude_kws if kw and kw in full)
                if hit_cnt >= 2:
                    continue

            # 검색어에 포함된 단어를 룰베이스에서 제외
            query_words = [w for w in item.get("검색어", "").split() if w not in DAISO_VARIANTS]
            sentiment, score, reason = ensemble_sentiment(r_batch[idx], full, threshold, exclude_words=query_words)

            date_str = item.get("날짜","") if src == "유튜브" else (
                lambda dt: dt.strftime("%Y-%m-%d") if dt else ""
            )(parse_date(item))

            prod_code = extract_product_code(full)
            prod_name = match_product_name(prod_code)
            subcategory = extract_subcategory(full)

            # 품번 매칭 실패 시 품명 역매칭으로 보완
            if not prod_code or not prod_name:
                name_match = match_by_product_name(full)
                if name_match["품명"]:
                    if not prod_code:
                        prod_code = name_match["품번"]
                    if not prod_name:
                        prod_name = name_match["품명"]
                    if not subcategory:
                        subcategory = name_match["소분류"]

            price_mention = extract_price(full) if src != "유튜브" else ""

            results.append({
                "출처":    src,
                "검색어":  item.get("검색어",""),
                "소분류":  subcategory,
                "품번":    prod_code,
                "품명":    prod_name,
                "가격언급": price_mention,
                "title":  title,
                "link":   item.get("link",""),
                "날짜":   date_str,
                "감성":   sentiment,
                "확신도": score,
                "channel": item.get("channel","") or item.get("cafename",""),
                "views":   item.get("views",""),
                "likes":   item.get("likes",""),
                "comments":item.get("comments",""),
                "video_id":item.get("video_id",""),
                "reason":  reason,
                "관련성":  item.get("_relevance", 0),
                "description": item.get("description", ""),  # AI 배치 라벨링 및 상세 확인용
            })

        done_so_far = min(batch_start + BATCH, total_f)
        progress_bar.progress(done_so_far / total_f)
        status_text.markdown(
            f'<span style="font-size:0.78rem;color:#718096;">AI 분석 중 {done_so_far} / {total_f} &nbsp;|&nbsp; KLUE-RoBERTa + 룰베이스</span>',
            unsafe_allow_html=True
        )

    progress_bar.empty(); status_text.empty()

    if EXCLUDE_SUBCATEGORIES:
        results = [r for r in results if not any(es in (r.get("소분류") or "") for es in EXCLUDE_SUBCATEGORIES)]

    # ── 분석 결과를 session_state에 저장 (sort/filter rerun 후에도 유지) ──
    st.session_state["analysis_results"] = results
    st.session_state["analysis_start_date"] = start_date
    st.session_state["analysis_end_date"] = end_date

# ── 결과가 session_state에 있으면 항상 탭 렌더링 ──
if "analysis_results" in st.session_state and st.session_state["analysis_results"]:
    results    = st.session_state["analysis_results"]
    start_date = st.session_state["analysis_start_date"]
    end_date   = st.session_state["analysis_end_date"]
    _tab_labels = ["대시보드", "블로그", "카페", "지식iN", "유튜브", "유튜브 댓글", "커뮤니티"]
    if st.session_state.get("admin_mode"):
        _tab_labels.append("관리자")
    _tabs = st.tabs(_tab_labels)
    tab_dash, tab_blog, tab_cafe, tab_kin, tab_yt, tab_yt_c, tab_ppomppu = (
        _tabs[0], _tabs[1], _tabs[2], _tabs[3], _tabs[4], _tabs[5], _tabs[6]
    )
    tab_admin = _tabs[7] if st.session_state.get("admin_mode") else None
    
    total = len(results)
    pos   = sum(1 for r in results if r["감성"]=="긍정")
    neg   = sum(1 for r in results if r["감성"]=="부정")
    neu   = sum(1 for r in results if r["감성"]=="중립")

    all_subs = []
    for r in results:
        if r.get("소분류"): all_subs.extend([s.strip() for s in r["소분류"].split(",") if s.strip()])
    sub_cnt = Counter(all_subs)
    
    all_codes = []
    for r in results:
        if r.get("품번"):
            for c in r["품번"].split(","):
                c = c.strip()
                if c: all_codes.append(f"{c} {r.get('품명','') }".strip())
    code_cnt = Counter(all_codes)
    
    date_neg = {}
    for r in results:
        if r["감성"] == "부정" and r.get("날짜"):
            month = r["날짜"][:7]
            date_neg[month] = date_neg.get(month, 0) + 1
    
    with tab_dash:
        st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">분석 요약</div>', unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        for col, cls, lbl, val, pct, ic_txt in [
            (c1,"total","전체 수집",  str(total), "100%",                                    "전체"),
            (c2,"pos",  "긍정",      str(pos),   f"{round(pos/total*100) if total else 0}%","긍정"),
            (c3,"neg",  "부정",      str(neg),   f"{round(neg/total*100) if total else 0}%","부정"),
            (c4,"neu",  "중립",      str(neu),   f"{round(neu/total*100) if total else 0}%","중립"),
        ]:
            with col:
                st.markdown(f"""
                <div class="metric-card {cls}">
                    <div class="metric-label">
                        <span class="metric-icon {cls}" style="color:#FFFFFF !important;">{ic_txt}</span>
                        {lbl}
                    </div>
                    <div class="metric-value">{val}</div>
                    <div class="metric-pct">{pct}</div>
                </div>""", unsafe_allow_html=True)
    
        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        d1, d2, d3 = st.columns(3)
        sub_u  = len(sub_cnt)
        code_u = len(set(r["품번"] for r in results if r.get("품번")))
        name_u = len(set(r["품명"] for r in results if r.get("품명")))
        for col, lbl, val in [(d1,"소분류 수",str(sub_u)),(d2,"품번 수",str(code_u)),(d3,"품명 수",str(name_u))]:
            with col:
                st.markdown(f"""
                <div class="card" style="text-align:center;padding:1rem 0.75rem;">
                    <div style="font-size:1.6rem;font-weight:700;color:#0066CC;font-family:'Inter',sans-serif;">{val}</div>
                    <div style="font-size:0.72rem;color:#718096;margin-top:0.2rem;font-weight:500;">{lbl}</div>
                </div>""", unsafe_allow_html=True)
    
        date_pos_week = {}
        date_neg_week = {}
        for r in results:
            if r.get("날짜") and len(r["날짜"]) >= 10:
                try:
                    dt = datetime.strptime(r["날짜"][:10], "%Y-%m-%d")
                    # 월요일 기준 주간 시작일
                    week_start = dt - pd.Timedelta(days=dt.weekday())
                    week_label = week_start.strftime("%m/%d") + "~"
                    if r["감성"] == "긍정":
                        date_pos_week[week_label] = date_pos_week.get(week_label, 0) + 1
                    elif r["감성"] == "부정":
                        date_neg_week[week_label] = date_neg_week.get(week_label, 0) + 1
                except:
                    pass

        all_weeks = sorted(set(list(date_pos_week.keys()) + list(date_neg_week.keys())))
        if all_weeks:
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">주간별 긍정 · 부정 추이</div>', unsafe_allow_html=True)
            chart_data = []
            for w in all_weeks:
                chart_data.append({"주간": w, "건수": date_pos_week.get(w, 0), "감성": "긍정"})
                chart_data.append({"주간": w, "건수": date_neg_week.get(w, 0), "감성": "부정"})
            chart_df = pd.DataFrame(chart_data)
            chart = (
                alt.Chart(chart_df)
                .mark_line(point=True, strokeWidth=2.5)
                .encode(
                    x=alt.X("주간:O", axis=alt.Axis(title="", labelAngle=-45, labelFontSize=10)),
                    y=alt.Y("건수:Q", axis=alt.Axis(title="건수", titleFontSize=11)),
                    color=alt.Color("감성:N", scale=alt.Scale(domain=["긍정","부정"], range=["#16A34A","#DC2626"]), legend=alt.Legend(title=None)),
                    tooltip=[alt.Tooltip("주간:O", title="주간"), alt.Tooltip("감성:N", title="감성"), alt.Tooltip("건수:Q", title="건수")]
                )
                .properties(height=220)
                .configure_view(strokeWidth=0)
                .configure_axis(grid=False, domain=False)
            )
            st.altair_chart(chart, use_container_width=True)

        # ── 부정 리뷰 워드클라우드 ──
        neg_all_text = " ".join(
            (r.get("title", "") + " " + r.get("description", "") + " " + (r.get("reason") or ""))
            for r in results if r.get("감성") == "부정"
        )
        if neg_all_text.strip():
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.5rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">부정 리뷰 핵심 단어</div>', unsafe_allow_html=True)
            with st.container():
                st.markdown('<div class="card" style="padding:0.75rem;">', unsafe_allow_html=True)
                render_wordcloud(neg_all_text, height=260, max_words=80, colormap="Reds")
                st.markdown('</div>', unsafe_allow_html=True)

        # ── 상품별 이슈 워드클라우드 (goldset에 저장된 issue 텍스트 기반) ──
        try:
            gs = load_goldset_from_sheet()
            issue_texts = " ".join(
                (r.get("issue") or "") for r in gs
                if (r.get("label") in ("부정", "긍정") and (r.get("issue") or "").strip())
            )
            if issue_texts.strip():
                st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.5rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">검토 목록 이슈 워드클라우드 <span style="color:#A0AEC0;font-weight:500;">· 관리자 라벨링 기반</span></div>', unsafe_allow_html=True)
                st.markdown('<div class="card" style="padding:0.75rem;">', unsafe_allow_html=True)
                render_wordcloud(issue_texts, height=220, max_words=60, colormap="Blues")
                st.markdown('</div>', unsafe_allow_html=True)
        except Exception:
            pass
    
        col_top1, col_top2 = st.columns(2)
        with col_top1:
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 0.75rem;">카테고리 상위 10</div>', unsafe_allow_html=True)
            html = ""
            for rank, (name, count) in enumerate(sub_cnt.most_common(10), 1):
                cls = "r1" if rank == 1 else ""
                html += f'<div class="top-item"><div class="top-rank {cls}" style="color:{"#FFFFFF" if rank==1 else "var(--primary)"};">{rank}</div><div class="top-name">{name}</div><div class="top-count">{count}건</div></div>'
            empty_sub_html = "<span style='color:#718096;font-size:0.82rem;'>소분류 데이터 없음</span>"
            st.markdown(f'<div class="card">{html or empty_sub_html}</div>', unsafe_allow_html=True)
    
        with col_top2:
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 0.75rem;">주요 품번 · 품명 상위 10</div>', unsafe_allow_html=True)
            html2 = ""
            for rank, (name, count) in enumerate(code_cnt.most_common(10), 1):
                cls = "r1" if rank == 1 else ""
                html2 += f'<div class="top-item"><div class="top-rank {cls}" style="color:{"#FFFFFF" if rank==1 else "var(--primary)"};">{rank}</div><div class="top-name">{name}</div><div class="top-count">{count}건</div></div>'
            empty_code_html = "<span style='color:#718096;font-size:0.82rem;'>품번 데이터 없음</span>"
            st.markdown(f'<div class="card">{html2 or empty_code_html}</div>', unsafe_allow_html=True)
    
        st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">주요 부정 글</div>', unsafe_allow_html=True)
        neg_results = [r for r in results if r["감성"] == "부정"]
        if neg_results:
            for r in neg_results[:20]:
                _b    = SENT_BADGE.get(r["감성"], "")
                _sub  = ('<span class="badge-sub">카테고리 · ' + r["소분류"] + '</span>') if r.get("소분류") else ""
                _code = ('<span class="badge-sub">품번 · ' + r["품번"] + '</span>') if r.get("품번") else ""
                _name = ('<span>품명 · ' + r["품명"] + '</span>') if r.get("품명") else ""
                _badge = '<span class="' + _b + '">' + r["감성"] + ' ' + fmt_score(r["확신도"]) + '</span>'
                _reason = ('<div style="font-size:0.75rem;color:#4A5568;margin-top:0.4rem;padding-left:0.7rem;border-left:2px solid #CBD5E1;font-family:\'IBM Plex Sans KR\',sans-serif;"><span style="font-size:0.62rem;color:#7A879E;font-family:\'IBM Plex Mono\',monospace;text-transform:uppercase;letter-spacing:0.05em;">근거 </span>' + r["reason"] + '</div>') if r.get("reason") else ""
                _title = r["title"] or "(제목 없음)"
                _html  = (
                    '<div class="result-card">'
                    '<div class="result-title">'
                    '<a href="' + r["link"] + '" target="_blank" style="color:#1A202C;text-decoration:none;">' + _title + '</a>'
                    '</div>'
                    '<div class="result-meta">'
                    '<span>출처 · ' + r["출처"] + '</span>'
                    '<span>검색어 · ' + r["검색어"] + '</span>'
                    '<span>날짜 · ' + r["날짜"] + '</span>'
                    + _sub + _code + _name + _badge +
                    '</div>'
                    + _reason +
                    '</div>'
                )
                st.markdown(_html, unsafe_allow_html=True)
        else:
            st.info("부정으로 분류된 글이 없습니다.")
    
        st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:1.5rem 0 0.75rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">결과 다운로드</div>', unsafe_allow_html=True)
        dl1, dl2 = st.columns(2)
        with dl1:
            buf = create_excel(results, start_date, end_date)
            st.download_button("EXCEL 다운로드", buf,
                f"ISSUE_{start_date}_{end_date}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with dl2:
            csv = pd.DataFrame(results).to_csv(index=False, encoding="utf-8-sig")
            st.download_button("CSV 다운로드", csv.encode("utf-8-sig"),
                f"ISSUE_{start_date}_{end_date}.csv", "text/csv", use_container_width=True)
    
    
    with tab_blog:
        render_detail_tab([r for r in results if r["출처"]=="블로그"], "블로그", start_date, end_date)
    
    with tab_cafe:
        render_detail_tab([r for r in results if r["출처"]=="카페"], "카페", start_date, end_date)
    
    with tab_kin:
        render_detail_tab([r for r in results if r["출처"]=="지식iN"], "지식iN", start_date, end_date)
    
    with tab_yt:
        render_detail_tab([r for r in results if r["출처"]=="유튜브"], "유튜브", start_date, end_date)
    
    with tab_yt_c:
        render_detail_tab([r for r in results if r["출처"]=="유튜브 댓글"], "유튜브 댓글", start_date, end_date)
    
    with tab_ppomppu:
        render_detail_tab([r for r in results if r["출처"]=="커뮤니티"], "커뮤니티", start_date, end_date)
    
    # ============================
    # 관리자 탭 (탭 내부에 배치)
    # ============================
    if tab_admin is not None:
        with tab_admin:
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 1rem;padding-bottom:0.4rem;border-bottom:1px solid #E1E7F0;">관리자 · 단어 / 링크 관리</div>', unsafe_allow_html=True)

            # ── Gemini 사용량 표시 ──
            _gm_used = _gemini_usage_today()
            _gm_pct = round(_gm_used / max(GEMINI_DAILY_LIMIT, 1) * 100, 1)
            _gm_color = "#DC2626" if _gm_used >= GEMINI_DAILY_LIMIT else ("#B45309" if _gm_pct >= 80 else "#16A34A")
            _gm_status = "한도 도달" if _gm_used >= GEMINI_DAILY_LIMIT else ("임박" if _gm_pct >= 80 else "정상")
            _gm_enabled_txt = "ON" if GEMINI_ENABLED else "OFF (secrets: GEMINI_ENABLED=false)"
            st.markdown(f"""
            <div class="card" style="padding:0.75rem 1rem;margin-bottom:0.75rem;">
                <div style="display:flex;justify-content:space-between;align-items:center;">
                    <div>
                        <div style="font-size:0.72rem;color:#7A879E;text-transform:uppercase;letter-spacing:0.05em;font-family:'IBM Plex Mono',monospace;">AI 오늘 사용량</div>
                        <div style="font-family:monospace;font-size:1.4rem;font-weight:700;color:{_gm_color};">{_gm_used} / {GEMINI_DAILY_LIMIT}회 · {_gm_pct}%</div>
                    </div>
                    <div style="text-align:right;">
                        <div style="font-size:0.9rem;font-weight:700;color:{_gm_color};">{_gm_status}</div>
                        <div style="font-size:0.7rem;color:#7A879E;font-family:monospace;">스위치: {_gm_enabled_txt}</div>
                    </div>
                </div>
                <div style="margin-top:0.6rem;height:6px;background:#E2E8F0;border-radius:3px;overflow:hidden;">
                    <div style="height:100%;width:{min(_gm_pct,100)}%;background:{_gm_color};"></div>
                </div>
                <div style="font-size:0.68rem;color:#7A879E;margin-top:0.5rem;">
                    ※ 이 카운터는 앱 세션 기준입니다. 앱이 재시작되면 리셋되지만 Google 서버의 실제 한도(일 1,500회)는 유지됩니다.
                </div>
            </div>
            """, unsafe_allow_html=True)

            sheet_kw = load_keywords_from_sheet()

            # ── 키워드 등록 (상단) ──
            st.markdown('<div style="font-family:\'IBM Plex Mono\',monospace;font-size:0.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 0.75rem;">단어 등록</div>', unsafe_allow_html=True)

            reg1, reg2 = st.columns(2)
            with reg1:
                _kw_type_map = {"제외 키워드": "exclude", "부정 키워드": "neg", "긍정 키워드": "pos", "홍보성 멘트": "promo"}
                kw_type_label = st.selectbox("유형 선택", list(_kw_type_map.keys()), key="admin_kw_type2")
                kw_type = _kw_type_map[kw_type_label]
                new_kw = st.text_input("키워드 입력", key="admin_new_kw2", placeholder="추가할 키워드 또는 문구 입력")
                add_col, del_col = st.columns(2)
                with add_col:
                    if st.button("추가", key="admin_add_kw2", use_container_width=True) and new_kw.strip():
                        existing = load_keywords_from_sheet().get(kw_type, [])
                        if new_kw.strip() in existing:
                            st.warning("이미 등록된 단어입니다.")
                        else:
                            append_keyword_to_sheet(kw_type, new_kw.strip())
                            if kw_type == "exclude":
                                st.session_state["admin_exclude_kws"].append(new_kw.strip())
                            st.success(f"[{kw_type_label}] '{new_kw.strip()}' 추가 완료")
                            st.rerun()
                with del_col:
                    if st.button("삭제", key="admin_del_kw2", use_container_width=True) and new_kw.strip():
                        if delete_keyword_from_sheet(kw_type, new_kw.strip()):
                            st.success(f"[{kw_type_label}] '{new_kw.strip()}' 삭제 완료")
                            st.rerun()
                        else:
                            st.warning("해당 단어를 찾을 수 없습니다.")

            with reg2:
                new_url = st.text_input("제외 URL 입력", key="admin_new_url", placeholder="https://...")
                url_reason = st.text_input("제외 사유", key="admin_url_reason", placeholder="(선택) 사유 입력")
                if st.button("링크 시트에 추가", key="admin_add_url", use_container_width=True) and new_url.strip():
                    if new_url.strip() in EXCLUDED_URLS_FROM_SHEET:
                        st.warning("이미 등록된 링크입니다.")
                    else:
                        append_excluded_url_to_sheet(new_url.strip(), url_reason.strip() or "관리자 수동 제외")
                        st.success("링크 제외 등록 완료")
                        st.rerun()

            st.markdown("---")

            # ── 뱃지 렌더링 함수 ──
            def _render_badges(title, icon_str, kw_list, bg_color, text_color):
                badges = "".join(
                    f'<span style="display:inline-block;font-size:0.78rem;background:{bg_color};color:{text_color};padding:3px 10px;border-radius:20px;margin:3px 4px;font-weight:500;">{kw}</span>'
                    for kw in kw_list
                ) if kw_list else '<span style="font-size:0.78rem;color:#A0AEC0;">등록된 항목 없음</span>'
                st.markdown(f'''
                <div class="card" style="margin-bottom:0.75rem;">
                    <div style="font-size:0.82rem;font-weight:700;color:{text_color};margin-bottom:0.5rem;">{icon_str} {title} ({len(kw_list)}건)</div>
                    <div style="display:flex;flex-wrap:wrap;gap:2px;">{badges}</div>
                </div>''', unsafe_allow_html=True)

            # ── 각 카테고리 뱃지 표시 ──
            _render_badges("광고 제외 단어", "", sheet_kw.get("exclude", []), "#FEF2F2", "#DC2626")
            _render_badges("부정 단어", "", sheet_kw.get("neg", []), "#FEF2F2", "#DC2626")
            _render_badges("긍정 단어", "", sheet_kw.get("pos", []), "#F0FDF4", "#16A34A")
            _render_badges("광고성 문구", "", sheet_kw.get("promo", []), "#FEFCE8", "#CA8A04")

            # ── 제외 URL (접기) ──
            with st.expander(f"제외한 링크 ({len(EXCLUDED_URLS_FROM_SHEET)}건)", expanded=False):
                url_list = list(EXCLUDED_URLS_FROM_SHEET)[:50]
                if url_list:
                    url_badges = "".join(
                        f'<span style="display:inline-block;font-size:0.7rem;background:#F1F5F9;color:#475569;padding:3px 8px;border-radius:4px;margin:3px 4px;word-break:break-all;max-width:100%;">{u}</span>'
                        for u in url_list
                    )
                    st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:2px;">{url_badges}</div>', unsafe_allow_html=True)
                else:
                    st.caption("등록된 항목 없음")

            st.markdown("---")
            _render_admin_analytics(results)

st.markdown("""
<div style="text-align:center;padding:2rem 0 1rem;border-top:1px solid #E2E8F0;margin-top:2rem;">
    <span style="font-size:0.75rem;color:#A0AEC0;">DAISO SNS ISSUE FINDER · KLUE-RoBERTa + 룰베이스 앙상블 · Created by 데이터분석팀</span>
</div>
""", unsafe_allow_html=True)
